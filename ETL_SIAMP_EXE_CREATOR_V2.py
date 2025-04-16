#!/usr/bin/env python
# -*- coding: utf-8 -*-

from gooey import Gooey, GooeyParser
import pandas as pd
import os
import re
import glob
from time import sleep
import warnings
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import sys
import io
import requests

# Chemin absolu vers l'icône (sera utilisée dans l'interface et pour l'exécutable)
ICON_PATH = "C:/Users/elias/OneDrive/Documents/PROFESSIONNEL/SIAMP/SUJETS/SUJET - 1 (Gestion C.A filiales)/Dev/siamp_icon.ico"

# Fichier de configuration pour sauvegarder la clé API
CONFIG_FILE = "siamp_api_key.cfg"
***REMOVED***

# Console UTF-8
if sys.stdout and hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


def load_api_key():
    """Charge la clé API depuis le fichier de configuration si disponible, sinon retourne la clé par défaut."""
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                key = f.read().strip()
                if key:
                    return key
        except Exception as e:
            print(f"[ERREUR] Impossible de lire le fichier de configuration API : {e}")
    return DEFAULT_API_KEY


def save_api_key(new_key):
    """Enregistre la clé API dans le fichier de configuration."""
    try:
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            f.write(new_key.strip())
        print("[INFO] Nouvelle clé API sauvegardée.")
    except Exception as e:
        print(f"[ERREUR] Impossible de sauvegarder la clé API : {e}")


def get_live_conversion_rates(api_key):
    """
    Récupère les taux de conversion via currencyapi.net en utilisant la clé API fournie.
    En cas d'erreur, lève l'exception afin d'informer l'utilisateur.
    """
    url = "https://currencyapi.net/api/v1/rates"
    params = {"key": api_key}
    response = requests.get(url, params=params)
    response.raise_for_status()
    data = response.json()

    if not data.get("valid", False):
        raise ValueError("Clé API invalide ou réponse incorrecte.")
    rates = data.get("rates", {})
    if "USD" not in rates:
        raise ValueError("USD manquant dans les taux retournés.")
    rates_eur_base = {}
    for code, taux in rates.items():
        try:
            taux = float(taux)
            if taux != 0:
                rates_eur_base[code.upper()] = taux
        except:
            continue
    rates_eur_base["EUR"] = 1.0

    print("[INFO] ✅ Taux de conversion vers EUR calculés correctement via API :")
    for k, v in rates_eur_base.items():
        print(f"  → {k} = {round(v, 6)} €")
    return rates_eur_base


@Gooey(
    program_name="ETL SIAMP",
    program_icon=ICON_PATH,
    header_logo=ICON_PATH,  # L'icône dans le header
    header_show_title=True,
    header_height=100,
    use_cmd_args=True
)
def main():
    # Récupère la clé API sauvegardée ou utilise la valeur par défaut
    api_key_saved = load_api_key()

    parser = GooeyParser(description="Fusionnez plusieurs fichiers Excel contenant des feuilles nommées Turnover")
    
    # Groupe pour le choix du mode
    mode_group = parser.add_argument_group("Mode de gestion des taux")
    mode_group.add_argument(
        "--mode",
        widget="Dropdown",
        choices=["API", "Manuel"],
        default="API",
        help="Sélectionnez le mode pour gérer les taux de change :\n"
             "API : Récupère les taux via currencyapi.net (clée API affichée ci-dessous).\n"
             "Manuel : Utilise uniquement des taux que vous renseignerez manuellement."
    )
    
    # Groupe pour les fichiers et le chemin de sortie
    input_group = parser.add_argument_group("Fichiers et sortie")
    input_group.add_argument(
        "--fichiers",
        widget="MultiFileChooser",
        nargs='+',
        required=True,
        help="Sélectionnez les fichiers Excel à fusionner (ex: fichiers_excel/*.xlsx)"
    )
    input_group.add_argument(
        "--chemin_sortie",
        widget="FileSaver",
        required=True,
        help="Chemin du fichier Excel final"
    )
    
    # Groupe pour la configuration en mode API
    api_group = parser.add_argument_group("Configuration API (currencyapi.net)")
    api_group.add_argument(
        "--api_key",
        widget="TextField",
        required=False,
        default=api_key_saved,
        help="Clé API pour récupérer les taux via currencyapi.net.\n"
             "Saisissez une nouvelle clé si l'ancienne a expiré. (La clé sera sauvegardée.)"
    )
    
    # Groupe pour la configuration en mode Manuel
    manuel_group = parser.add_argument_group("Configuration Manuel")
    manuel_group.add_argument(
        "--taux_manuels",
        widget="TextField",
        required=False,
        help="Taux manuels au format 'USD=0.93,GBP=1.15'.\n"
             "Utilisez ce champ en mode Manuel ou pour compléter le mode API."
    )
    
    args = parser.parse_args()

    # Sauvegarde la clé API si elle a été modifiée
    if args.api_key.strip() != api_key_saved:
        save_api_key(args.api_key.strip())
    api_key = args.api_key.strip()

    def parse_taux_manuels(taux_str):
        taux_dict = {}
        if not taux_str:
            return taux_dict
        try:
            for paire in taux_str.split(","):
                code, val = paire.strip().split("=")
                taux_dict[code.strip().upper()] = float(val.strip())
        except Exception as e:
            print(f"[ERREUR] Format invalide pour les taux manuels : {e}")
        return taux_dict

    # Récupération des fichiers (extension .xlsx) en gérant les jokers
    fichiers = []
    for path in args.fichiers:
        fichiers.extend(glob.glob(path))
    fichiers = [f for f in fichiers if f.endswith('.xlsx') and not os.path.basename(f).startswith('~$')]

    # Préparation du chemin de sortie
    chemin_final = args.chemin_sortie
    dossier_sortie = os.path.dirname(chemin_final)
    if not dossier_sortie:
        dossier_sortie = "."
        chemin_final = os.path.join(dossier_sortie, chemin_final)
    os.makedirs(dossier_sortie, exist_ok=True)
    if not chemin_final.lower().endswith(".xlsx"):
        chemin_final += ".xlsx"

    # Initialisation de la variable de taux de conversion
    taux_conversion = {}
    if args.mode.lower() == "api":
        try:
            print("[INFO] Mode API sélectionné. Tentative de récupération des taux via currencyapi.net...")
            taux_conversion = get_live_conversion_rates(api_key)
            print("[INFO] ✅ Clé API fonctionnelle.")
        except Exception as error:
            error_message = str(error)
            print(f"[ERREUR] La clé API ne fonctionne pas : {error_message}")
            print("[INFO] Veuillez vérifier votre clé ou passer en mode Manuel.")
            # Fallback sur des taux par défaut
            taux_conversion = {
                "EUR": 1.0,
                "USD": 0.93,
                "GBP": 1.15,
                "EGP": 0.03,
                "CHF": 1.04,
                "AED": 0.25,
                "JPY": 0.0062
            }
    else:
        print("[INFO] Mode Manuel sélectionné : utilisation uniquement des taux manuels.")
    
    # Mise à jour avec les taux manuels (ils ont la priorité sur ceux récupérés via API)
    taux_manuels = parse_taux_manuels(args.taux_manuels)
    if taux_manuels:
        print("[INFO] 🔧 Taux manuels fournis :")
        for k, v in taux_manuels.items():
            print(f"  → {k} = {v} € (prioritaire)")
        taux_conversion.update(taux_manuels)
    elif args.mode.lower() == "manuel" and not taux_manuels:
        print("[AVERTISSEMENT] Aucun taux manuel fourni, utilisation des taux fixes par défaut.")
        taux_conversion = {
            "EUR": 1.0,
            "USD": 0.93,
            "GBP": 1.15,
            "EGP": 0.03,
            "CHF": 1.04,
            "AED": 0.25,
            "JPY": 0.0062
        }

    # Expression du pattern pour identifier les feuilles "Turnover"
    pattern_turnover = re.compile(r"^Turnover$|^TURNOVER$|^Turnover\s+[A-Z][a-z]{2}\s+\d{1,2}$")
    dfs = []
    total = len(fichiers)

    print("Début de la fusion des fichiers...\n")

    for i, fichier in enumerate(fichiers):
        print(f"🔍 Analyse du fichier : {os.path.basename(fichier)}", flush=True)
        try:
            xls = pd.ExcelFile(fichier, engine="openpyxl")
            feuilles = [s for s in xls.sheet_names if pattern_turnover.match(s)]
            if not feuilles:
                print(f"⚠️ Aucune feuille Turnover détectée dans {os.path.basename(fichier)}. Vérifiez son format.", flush=True)
                continue

            for feuille in feuilles:
                print(f"✅ Feuille trouvée : {feuille} ({os.path.basename(fichier)})", flush=True)
                df = xls.parse(feuille, usecols="A:O")
                df.dropna(axis=1, how="all", inplace=True)
                df.columns = [col.strip().upper() for col in df.columns]
                rename_dict = {}
                for col in df.columns:
                    if col == "CURRENCY":
                        rename_dict[col] = "Currency"
                    elif col == "TURNOVER":
                        rename_dict[col] = "TURNOVER"
                    elif col == "CUSTOMER NAME":
                        rename_dict[col] = "Customer Name"
                    elif col == "CUSTOMER":
                        rename_dict[col] = "Customer Name"
                df.rename(columns=rename_dict, inplace=True)

                turnover_col = "TURNOVER" if "TURNOVER" in df.columns else None
                quantity_col = "QUANTITY" if "QUANTITY" in df.columns else None
                if turnover_col or quantity_col:
                    df = df[~(df.get(turnover_col).isna() & df.get(quantity_col).isna())]

                if "Currency" in df.columns:
                    df["Currency"] = df["Currency"].fillna(method="ffill")

                if "Currency" in df.columns and "TURNOVER" in df.columns:
                    def conversion(row):
                        devise = str(row["Currency"]).strip().upper()
                        montant = row["TURNOVER"]
                        if pd.isna(montant):
                            return None
                        if not devise or devise == "NAN":
                            print(f"[AVERTISSEMENT] Aucune devise indiquée pour une ligne de {os.path.basename(fichier)}", flush=True)
                            return None
                        taux = taux_conversion.get(devise)
                        if taux:
                            print(f"[DEBUG] {montant} {devise} → {round(montant / taux, 2)} EUR via taux {taux}")
                            return round(montant / taux, 2)
                        else:
                            print(f"[AVERTISSEMENT] Devise inconnue '{devise}' dans {os.path.basename(fichier)}", flush=True)
                            return None

                    df.insert(df.columns.get_loc("TURNOVER") + 1, "C.A en €", df.apply(conversion, axis=1))

                colonnes_a_verifier = [col for col in ["TURNOVER", "QUANTITY"] if col in df.columns]
                if colonnes_a_verifier:
                    masque = pd.Series(True, index=df.index)
                    for col in colonnes_a_verifier:
                        masque &= pd.to_numeric(df[col], errors="coerce").notna()
                    lignes_supprimees = (~masque).sum()
                    if lignes_supprimees > 0:
                        print(f"[INFO] {lignes_supprimees} ligne(s) supprimée(s) pour valeurs non numériques dans {', '.join(colonnes_a_verifier)}.", flush=True)
                    df = df[masque]

                df["NomFichier"] = os.path.basename(fichier)
                df["Feuille"] = feuille
                dfs.append(df)
                print(f"[OK] Feuille ajoutée : {feuille} ({os.path.basename(fichier)})")

        except Exception as e:
            print(f"[ERREUR] Problème avec {fichier} : {e}")

        pourcentage = int(((i + 1) / total) * 100)
        print(f"PROGRESS: {pourcentage}%", flush=True)
        if pourcentage == 100:
            print("\n⏳ Les données sont entièrement chargées. Veuillez patienter pendant la finalisation du fichier Excel (ne fermez pas l'application)...", flush=True)
        sleep(0.1)

    if not dfs:
        print("\nAucun fichier ou feuille valide détecté. Arrêt.")
        return

    fusion = pd.concat(dfs, ignore_index=True)
    fusion.to_excel(chemin_final, index=False)

    wb = load_workbook(chemin_final)
    ws = wb.active
    max_col = ws.max_column
    max_row = ws.max_row
    end_col_letter = get_column_letter(max_col)
    table_range = f"A1:{end_col_letter}{max_row}"

    table = Table(displayName="FusionTable", ref=table_range)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)

    for col in ws.iter_cols(min_row=2, max_row=max_row):
        header_cell = ws[f"{col[0].column_letter}1"]
        if header_cell.value == "C.A en €":
            for cell in col:
                cell.number_format = u"#,##0.00\u00a0€"

    wb.save(chemin_final)

    recap = "\n=== ✅ FUSION COMPLÉTÉE AVEC SUCCÈS ===\n"
    recap += f"📄 Fichier généré : {chemin_final}\n"
    recap += "\nMerci d’avoir utilisé l’outil ETL SIAMP. 🚀\n"
    print(recap, flush=True)


if __name__ == '__main__':
    main()
