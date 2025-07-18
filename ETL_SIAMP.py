#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
ETL_SIAMP.py – fusion & enrichissement Turnover

• Récupère les taux historiques si votre plan le permet (/historical),
  sinon bascule automatiquement sur le temps réel (/rates).
• Ajoute VARIABLE COSTS (CD+FSD) et COGS (PRU) quelle que soit l'écriture.
• Maintient le calcul « C.A en € ».Fdate
• Réordonne les colonnes métier.
"""
from __future__ import annotations
import argparse
import glob
import io
import os
import re
import sys
import warnings
import configparser
import traceback
from time import sleep
from typing import Any
import xml.etree.ElementTree as ET
from datetime import datetime
import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import subprocess

# ------------------------------------------------------------------ console UTF‑8
if sys.stdout and hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


# ------------------------------------------------------------------ taux de change
import requests
import xml.etree.ElementTree as ET
from datetime import datetime

def get_ecb_rates(date: str | None = None, required_currencies: set[str] | None = None):
    print(f"[DEBUG] Appel get_ecb_rates(date={date})", flush=True)
    if date:
        url = "https://www.ecb.europa.eu/stats/eurofxref/eurofxref-hist.xml"
    else:
        url = "https://www.ecb.europa.eu/stats/eurofxref/eurofxref-daily.xml"

    try:
        response = requests.get(url)
        print(f"[INFO] 📡 Requête vers {url}", flush=True)
        print(f"[INFO] ✅ Statut : {response.status_code}", flush=True)
        response.raise_for_status()

        root = ET.fromstring(response.content)
        ns = {'ns': 'http://www.ecb.int/vocabulary/2002-08-01/eurofxref'}

        rates = {"EUR": 1.0}
        from datetime import datetime, timedelta

        if date:
            limit_date = (datetime.strptime(date, "%Y-%m-%d") - timedelta(days=60)).strftime("%Y-%m-%d")
            print(f"[INFO] 🔍 Recherche limitée aux taux entre {limit_date} et {date}", flush=True)



        dates = [cube.attrib["time"] for cube in root.findall(".//ns:Cube[@time]", ns)]
        if date:
            dates = sorted([d for d in dates if limit_date <= d <= date], reverse=True)
        else:
            dates = sorted(dates, reverse=True)


        rates_found = set(rates.keys())
        target_cube = None

        for d in dates:
            cube_d = root.find(f".//ns:Cube[@time='{d}']", ns)
            if cube_d is None:
                continue

            for cube in cube_d.findall("ns:Cube", ns):
                cur = cube.attrib["currency"]
                if cur not in rates:
                    rate = float(cube.attrib["rate"])
                    rates[cur] = rate
                    print(f"[INFO] ➕ Taux récupéré pour {cur} au {d} = {rate}", flush=True)
                    rates_found.add(cur)

            if required_currencies and required_currencies <= rates_found:
                print(f"[INFO] ✅ Tous les taux requis trouvés avant {d}", flush=True)
                break


        if date:
            # chercher le jour exact OU le plus proche avant
            dates = [cube.attrib["time"] for cube in root.findall(".//ns:Cube[@time]", ns)]
            print(f"[INFO] 📅 {len(dates)} dates trouvées dans l'historique ECB", flush=True)
            print(f"[INFO] 📅 Premières dates disponibles : {dates[:5]}", flush=True)
            dates.sort(reverse=True)
            target_date = None
            for d in dates:
                if d <= date:
                    target_date = d
                    break

            if not target_date:
                raise ValueError(f"Aucun taux trouvé avant la date {date}")

            target_cube = root.find(f".//ns:Cube[@time='{target_date}']", ns)
            if target_date != date:
                print(f"[INFO] ⚠️ Pas de taux pour {date}, utilisation de {target_date} à la place", flush=True)
            else:
                print(f"[INFO] ✅ Taux trouvés pour la date exacte : {target_date}", flush=True)


            if target_date != date:
                print(f"[INFO] ⚠ Aucun taux pour {date}, substitution par {target_date}", flush=True)
        else:
            # date non spécifiée : dernier taux connu
            cubes = root.findall(".//Cube[@time]")
            if not cubes:
                raise ValueError("Pas de données de taux trouvées")
            target_cube = cubes[0]
            target_date = target_cube.attrib["time"]
        
        print("[INFO] 🔎 Récupération des taux de conversion :", flush=True)
        for cube in target_cube.findall("ns:Cube", ns):
            currency = cube.attrib["currency"]
            raw_rate = float(cube.attrib["rate"])
            print(f"  → {currency} = {raw_rate}", flush=True)
            if raw_rate != 0:
                rates[currency] = raw_rate
        rates["EUR"] = 1.0

        if required_currencies:
            missing = required_currencies - rates_found
            if missing:
                print(f"[WARN] ❌ Aucun taux trouvé pour {sorted(missing)} dans les 60 derniers jours.", flush=True)
                print(f"[SUGGESTION] ✍️ Veuillez les ajouter manuellement dans l'interface ou en ligne de commande.", flush=True)



        print(f"[INFO] Taux ECB récupérés au {date}", flush=True)
        for k, v in rates.items():
            print(f"  → {k} = {v}")
        return rates

    except Exception as e:
        print(f"[ERROR] Erreur récupération ECB : {e}", flush=True)
        print("[FALLBACK] 🛑 Repli sur taux locaux codés en dur", flush=True)
        return {
        "EUR":1.0, "USD":0.93, "GBP":1.15,
        "EGP":0.03, "CHF":1.04, "AED":0.25, "JPY":0.0062
    }

# ------------------------------------------------------------------ CLI
def main():
    parser = argparse.ArgumentParser(description="Fusionnez plusieurs fichiers Excel Turnover")
    parser.add_argument("--fichiers",      nargs='+', required=True)
    parser.add_argument("--chemin_sortie", required=True)
    parser.add_argument("--taux_manuels",  help="USD=0.93,GBP=1.15", default=None)
    parser.add_argument("--date",          help="YYYY-MM-DD pour historique (premium)", default=None)
    parser.add_argument("--date_debut", help="Date début de la période à filtrer (YYYY-MM-DD)", default=None)
    parser.add_argument("--date_fin",   help="Date fin de la période à filtrer (YYYY-MM-DD)", default=None)
    parser.add_argument("--mois_selectionnes", help="Liste des mois à traiter, séparés par des virgules (ex: 2025-02,2025-03)", default=None)

    args = parser.parse_args()
    # ----------------------------------------- Charger les chemins des fichiers de référence
    def resource_path(relative_path):
        """Retourne le chemin absolu d'un fichier de ressources, compatible avec PyInstaller."""
        if hasattr(sys, "_MEIPASS"):
            return os.path.join(sys._MEIPASS, relative_path)  # exécutable PyInstaller
        return os.path.join(os.path.abspath("."), relative_path)  # mode normal
    
    CONFIG_REF_FILE = resource_path("mydata/ref_files.cfg")
    reference_file_path = None

    if os.path.exists(CONFIG_REF_FILE):
        config = configparser.ConfigParser()
        config.read(CONFIG_REF_FILE)
        refs = config['REFERENCES']
        reference_file_path = refs.get('reference_file', None)
    else:
        print("[WARN] ⚠️ Fichier de config 'ref_files.cfg' introuvable. Les colonnes de correspondance ne seront pas alimentées.")


    devises_detectées: set[str] = set()

    print(f"[DEBUG] 👋 Script lancé avec date = {args.date}", flush=True)

    # parse manuels
    manu: dict[str,float] = {}
    if args.taux_manuels:
        for part in args.taux_manuels.split(","):
            try:
                c,v = part.split("=")
                manu[c.strip().upper()] = float(v)
            except:
                print(f"[WARN] taux manuel ignoré: {part}", flush=True)

    # collecte fichiers
    files: list[str] = []
    for patt in args.fichiers:
        files.extend(glob.glob(patt))
    files = [f for f in files if f.lower().endswith(".xlsx")
             and not os.path.basename(f).startswith("~$")]
    if not files:
        sys.exit("Aucun fichier .xlsx trouvé.")

    out = args.chemin_sortie
    if not out.lower().endswith(".xlsx"):
        out += ".xlsx"
    os.makedirs(os.path.dirname(out) or ".", exist_ok=True)

    # patterns
    TURNOVER_SHEET = re.compile(r"^TURNOVER($|\s+[A-Z][a-z]{2}\s+\d{1,2}$)", re.I)
    VAR_PATTS  = [r"^CD\s*\+\s*FSD", r"^CD\+FSD", r"^VARIABLE\s*COSTS?"]
    COGS_PATTS = [r"^PRU", r"^COGS"]

    all_dfs: list[pd.DataFrame] = []
    fichiers_ignores = []  # Pour stocker les fichiers ignorés et leurs motifs
    total = len(files)
    for idx, path in enumerate(files, 1):
        print(f"[{idx}/{total}] {os.path.basename(path)}", flush=True)
        try:
            xls = pd.ExcelFile(path, engine="openpyxl")
            for sh in filter(TURNOVER_SHEET.match, xls.sheet_names):
                df = xls.parse(sh, usecols="A:Q")
                print(f"[AUDIT] Shape après lecture du fichier {os.path.basename(path)} / feuille {sh} : {df.shape}", flush=True)
                df.columns = [c.strip() for c in df.columns]
                print(f"[AUDIT] Shape après strip des colonnes : {df.shape}", flush=True)

                # renommage
                ren: dict[str,str] = {}
                for c in df.columns:
                    U = c.upper()
                    if any(re.match(p,U) for p in VAR_PATTS):
                        ren[c] = "VARIABLE COSTS"
                    elif any(re.match(p,U) for p in COGS_PATTS):
                        ren[c] = "COGS"
                    elif U=="TURNOVER":
                        ren[c] = "TURNOVER"
                    elif U=="CURRENCY":
                        ren[c] = "CURRENCY"
                    elif U in {"CUSTOMER","CUSTOMER NAME"}:
                        ren[c] = "CUSTOMER NAME"
                df.rename(columns=ren, inplace=True)
                print(f"[AUDIT] Shape après renommage des colonnes : {df.shape}", flush=True)

                # log var/cogs
                for nm in ("VARIABLE COSTS","COGS"):
                    if nm in df.columns:
                        n = df[nm].notna().sum()
                        print(f"       • {nm} détectée: {n} valeurs non-null", flush=True)

                df["NOMFICHIER"] = os.path.basename(path)
                df["FEUILLE"]     = sh
                df["SOURCE"] = f"MENSUEL_{datetime.now().strftime('%Y-%m-%d')}"
                print(f"[AUDIT] Shape après ajout des colonnes NOMFICHIER, FEUILLE, SOURCE : {df.shape}", flush=True)

                # ➤ Filtrer les lignes vides pour éviter de dépasser la limite Excel
                initial_rows = df.shape[0]
                # Supprimer les lignes où toutes les colonnes importantes sont vides
                important_cols = ["MONTH", "SIAMP UNIT", "SALE TYPE", "TYPE OF CANAL", "ENSEIGNE", "CUSTOMER NAME", "PRODUCT NAME", "TURNOVER"]
                available_cols = [col for col in important_cols if col in df.columns]
                
                if available_cols:
                    # Créer un masque pour les lignes non vides
                    non_empty_mask = df[available_cols].notna().any(axis=1)
                    df = df[non_empty_mask]
                    
                    removed_rows = initial_rows - df.shape[0]
                    if removed_rows > 0:
                        print(f"[WARN] ⚠️ {removed_rows} lignes vides supprimées de {os.path.basename(path)} pour éviter la limite Excel", flush=True)
                        if removed_rows > initial_rows * 0.5:  # Plus de 50% de lignes vides
                            print(f"[WARN] ⚠️ ATTENTION: {os.path.basename(path)} contient {removed_rows}/{initial_rows} lignes vides ({removed_rows/initial_rows*100:.1f}%)", flush=True)
                
                print(f"[AUDIT] Shape après filtrage des lignes vides : {df.shape}", flush=True)

                # Conversion explicite de la première colonne (MONTH) en datetime si possible
                if "MONTH" in df.columns:
                    try:
                        initial_month_rows = df.shape[0]
                        df["MONTH"] = pd.to_datetime(df["MONTH"], errors="coerce")
                        print(f"[AUDIT] Shape après conversion de MONTH en datetime : {df.shape}", flush=True)
                    except Exception as e:
                        print(f"[AUDIT] Shape après erreur conversion de MONTH : {df.shape}", flush=True)

                # Définir les deux formats stricts
                FORMATS = [
                    ["MONTH", "SIAMP UNIT", "SALE TYPE", "TYPE OF CANAL", "ENSEIGNE", "CUSTOMER NAME", "COMMERCIAL AREA", "SUR FAMILLE", "FAMILLE", "REFERENCE", "PRODUCT NAME", "QUANTITY", "TURNOVER", "CURRENCY", "COUNTRY", "VARIABLE COSTS", "COGS"],
                    ["MONTH", "SIAMP UNIT", "SALE TYPE", "TYPE OF CANAL", "ENSEIGNE", "CUSTOMER NAME", "COMMERCIAL AREA", "SUR FAMILLE", "FAMILLE", "REFERENCE", "PRODUCT NAME", "QUANTITY", "TURNOVER", "CURRENCY", "COUNTRY"]
                ]

                # VALIDATION STRICTE
                is_valid, motif, cols_manquantes, cols_sup = validate_strict_columns(df, os.path.basename(path), FORMATS, return_details=True)
                print(f"[AUDIT] Shape juste avant ajout à all_dfs : {df.shape}", flush=True)
                if is_valid:
                    all_dfs.append(df)
                    print(f"[AUDIT] Ajout de {os.path.basename(path)} au all_dfs. Taille actuelle de all_dfs: {len(all_dfs)}", flush=True)
                else:
                    fichiers_ignores.append({
                        'fichier': os.path.basename(path),
                        'motif': motif,
                        'colonnes_manquantes': cols_manquantes,
                        'colonnes_sup': cols_sup
                    })
                    print(f"\n❌ [IGNORÉ] {os.path.basename(path)} : Fichier non conforme, il ne sera pas fusionné.", flush=True)
                    if cols_manquantes:
                        print(f"   → Colonnes manquantes : {cols_manquantes}", flush=True)
                    if cols_sup:
                        print(f"   → Colonnes supplémentaires : {cols_sup}", flush=True)

        except Exception as e:
            print(f"[AUDIT] Erreur lors du traitement de {os.path.basename(path)} : {e}", flush=True)

        sleep(0.05)
        print(f"PROGRESS:{int(idx/total*100)}%", flush=True)

    if not all_dfs:
        print("[AUDIT] Aucun DataFrame valide dans all_dfs", flush=True)
        sys.exit("Aucune feuille valide trouvée.")
    
    initial_fusion_rows = sum(df.shape[0] for df in all_dfs)
    fusion = pd.concat(all_dfs, ignore_index=True)
    print(f"[AUDIT] Somme des lignes de all_dfs avant concat : {initial_fusion_rows}", flush=True)
    print(f"[AUDIT] Shape après pd.concat initial: {fusion.shape}. Total lignes attendues de all_dfs: {initial_fusion_rows}", flush=True)
    if fusion.shape[0] != initial_fusion_rows:
        print(f"[AUDIT] ATTENTION: pd.concat a modifié le nombre de lignes (de {initial_fusion_rows} à {fusion.shape[0]})", flush=True)


    # ➕ Convertir en majuscules (important)
    devises_detectées = {d.upper() for d in devises_detectées}

    # ✅ Maintenant que les devises sont détectées, on appelle la fonction
    rates = get_ecb_rates(args.date, required_currencies=devises_detectées)
    rates.update(manu)

    zone_affectation_df = None
    table_df = None

    if reference_file_path and os.path.exists(reference_file_path):
        try:
            table_df = pd.read_excel(reference_file_path, sheet_name="table", engine="openpyxl", dtype=str)
            print(f"[AUDIT] Shape de table_df après lecture : {table_df.shape}")
            print(f"[AUDIT] Colonnes présentes dans table_df : {list(table_df.columns)}")
            for col in ["REFERENCE V2", "REFERENCE", "ENSEIGNE V2", "ENSEIGNE", "CONCAT NAME"]:
                if col in table_df.columns:
                    non_null = table_df[col].notna().sum()
                    distinct = table_df[col].nunique(dropna=True)
                    print(f"[AUDIT] {col} : {non_null} valeurs non-nulles, {distinct} distinctes.")
                    print(f"[AUDIT] Exemples {col} : {table_df[col].dropna().unique()[:10]}")
            print(f"[AUDIT] ⚠️ Aucun drop_duplicates global n'est appliqué sur table_df. Toutes les lignes sont conservées.")
        except Exception as e:
            print(f"[AUDIT] Erreur chargement table : {e}")
            print(f"[AUDIT] Erreur chargement table : {e}")


    fusion = pd.concat(all_dfs, ignore_index=True)
    print(f"[AUDIT] Shape après pd.concat (juste avant nettoyage des strings) : {fusion.shape}", flush=True)

    # ➤ Nettoyage des chaînes de caractères : strip, upper, suppression des caractères invisibles
    def nettoyer_str(s):
        if pd.isna(s):
            return None
        if isinstance(s, str):
            s = s.strip().upper()
            s = re.sub(r'[^\x20-\x7E\u00A0-\uFFFF]', '', s)  # supprime caractères invisibles
            return s
        return s

    # Appliquer à toutes les colonnes de type objet (texte)
    for col in fusion.select_dtypes(include="object").columns:
        fusion[col] = fusion[col].apply(nettoyer_str)
    print(f"[AUDIT] Shape après nettoyage global des chaînes : {fusion.shape}", flush=True)
    print(f"[AUDIT] Nombre de doublons dans fusion après nettoyage : {fusion.duplicated().sum()}", flush=True)

    print(f"[DEBUG] 📌 Rates récupérés : {rates}", flush=True)
    currencies_in_file = set(fusion["CURRENCY"].dropna().unique())
    print(f"[DEBUG] 📌 Devises trouvées dans les fichiers : {currencies_in_file}", flush=True)
    missing_currencies = currencies_in_file - set(rates.keys())
    if missing_currencies:
        print(f"[ERROR] ❌ Aucune correspondance de taux pour les devises suivantes : {missing_currencies}", flush=True)
        print("         ➡️ Ajoutez-les dans les taux manuels ou vérifiez les données sources.", flush=True)
        sys.exit(1)
    else:
        print("[INFO] ✅ Tous les taux de conversion sont disponibles pour les devises présentes.", flush=True)


    # ---------------------------- ZONE AFFECTATION ----------------------------
    try:
        zone_affectation_df = pd.read_excel(
            reference_file_path,
            sheet_name="ZONE AFFECTATION",
            usecols="A,E",  # A = PAYS, E = Zone commerciale
            engine="openpyxl"
        )
        print(f"[AUDIT] Shape de zone_affectation_df : {zone_affectation_df.shape}", flush=True)
        print(f"[AUDIT] Nombre de doublons dans zone_affectation_df (PAYS) : {zone_affectation_df['PAYS'].duplicated().sum()}", flush=True)

        zone_affectation_df.columns = ["PAYS", "COMMERCIAL AREA"]
        fusion["COUNTRY"] = fusion["COUNTRY"].astype(str).str.strip().str.upper()
        zone_affectation_df["PAYS"] = zone_affectation_df["PAYS"].astype(str).str.strip().str.upper()
        
        initial_fusion_rows = fusion.shape[0]
        fusion = fusion.merge(zone_affectation_df, how="left", left_on="COUNTRY", right_on="PAYS")
        print(f"[AUDIT] Shape après fusion COMMERCIAL AREA : {fusion.shape}", flush=True)
        if fusion.shape[0] != initial_fusion_rows:
            print(f"[AUDIT] ATTENTION: Fusion COMMERCIAL AREA a modifié le nombre de lignes (de {initial_fusion_rows} à {fusion.shape[0]})", flush=True)

        fusion.drop(columns=["PAYS"], inplace=True)
        if "COMMERCIAL AREA_x" in fusion.columns and "COMMERCIAL AREA_y" in fusion.columns:
            fusion.drop(columns=["COMMERCIAL AREA_x"], inplace=True)
            fusion.rename(columns={"COMMERCIAL AREA_y": "COMMERCIAL AREA"}, inplace=True)
        elif "COMMERCIAL AREA_y" in fusion.columns:
            fusion.rename(columns={"COMMERCIAL AREA_y": "COMMERCIAL AREA"}, inplace=True)
        print(f"[INFO] ✅ Fusion COMMERCIAL AREA effectuée.")
    except Exception as e:
        print(f"[AUDIT] Erreur fusion ZONE AFFECTATION : {e}", flush=True)
        traceback.print_exc()
        print(f"[AUDIT] Erreur fusion ZONE AFFECTATION : {e}", flush=True)

    # ---------------------------- SURFAMILLE RET ----------------------------
    try:
        print(f"[AUDIT] Shape de table_df juste avant mapping Surfamille ret : {table_df.shape if table_df is not None else 'table_df=None'}")
        print("[INFO] 🔍 Nouveau mapping Surfamille ret (V2 puis fallback)...")
        # Fonction de normalisation forte
        def normalize_ref(ref):
            if pd.isna(ref):
                return ""
            return re.sub(r'[^A-Z0-9]', '', str(ref).upper())
        # Appliquer la normalisation sur les références d'origine
        fusion["REFERENCE_NORM"] = fusion["REFERENCE"].apply(normalize_ref)
        print(f"[DEBUG] Exemples REFERENCE_NORM fusion : {fusion['REFERENCE_NORM'].unique()[:10]}")
        if table_df is not None:
            # --- LOG: Stats sur les colonnes de mapping V2 ---
            if "REFERENCE V2" in table_df.columns and "Surfamille ret V2" in table_df.columns:
                table_df["REFERENCE_V2_NORM"] = table_df["REFERENCE V2"].apply(normalize_ref)
                table_df["Surfamille ret V2"] = table_df["Surfamille ret V2"].astype(str).str.strip().str.upper()
                print(f"[DEBUG] Exemples REFERENCE_V2_NORM table : {table_df['REFERENCE_V2_NORM'].unique()[:10]}")
                mapping_v2 = dict(zip(table_df["REFERENCE_V2_NORM"], table_df["Surfamille ret V2"]))
                fusion["Surfamille ret"] = fusion["REFERENCE_NORM"].map(mapping_v2)
                found_v2 = fusion["Surfamille ret"].notna().sum()
                print(f"[LOG] Après mapping V2 (normalisé) : {found_v2} correspondances trouvées sur {len(fusion)} lignes.")
                if found_v2 > 0:
                    print(f"[LOG] Exemples de valeurs enrichies (V2) : {fusion.loc[fusion['Surfamille ret'].notna(), ['REFERENCE','Surfamille ret']].head(5).to_dict(orient='records')}")
                not_found_v2 = fusion.loc[fusion["Surfamille ret"].isna(), "REFERENCE"].unique()[:10]
                print(f"[LOG] Exemples de REFERENCE non trouvées en V2 : {not_found_v2}")
            else:
                fusion["Surfamille ret"] = None
            # --- Fallback ancien mapping ---
            mask_vide = fusion["Surfamille ret"].isna()
            if mask_vide.any() and "REFERENCE" in table_df.columns and "Surfamille ret" in table_df.columns:
                table_df["REFERENCE_NORM"] = table_df["REFERENCE"].apply(normalize_ref)
                table_df["Surfamille ret"] = table_df["Surfamille ret"].astype(str).str.strip().str.upper()
                print(f"[DEBUG] Exemples REFERENCE_NORM table (fallback) : {table_df['REFERENCE_NORM'].unique()[:10]}")
                mapping_old = dict(zip(table_df["REFERENCE_NORM"], table_df["Surfamille ret"]))
                fusion.loc[mask_vide, "Surfamille ret"] = fusion.loc[mask_vide, "REFERENCE_NORM"].map(mapping_old)
                found_fallback = fusion["Surfamille ret"].notna().sum() - found_v2
                print(f"[LOG] Après fallback (normalisé) : {found_fallback} correspondances trouvées en plus.")
                if found_fallback > 0:
                    print(f"[LOG] Exemples de valeurs enrichies (fallback) : {fusion.loc[mask_vide & fusion['Surfamille ret'].notna(), ['REFERENCE','Surfamille ret']].head(5).to_dict(orient='records')}")
                not_found_final = fusion.loc[fusion["Surfamille ret"].isna(), "REFERENCE"].unique()[:10]
                print(f"[LOG] Exemples de REFERENCE toujours non trouvées : {not_found_final}")
            # Si toujours rien, laisser vide
            mask_vide = fusion["Surfamille ret"].isna()
            if mask_vide.any():
                print(f"[LOG] {mask_vide.sum()} lignes sans aucune correspondance pour Surfamille ret.")
                fusion.loc[mask_vide, "Surfamille ret"] = None
        else:
            fusion["Surfamille ret"] = None
        print(f"[INFO] ✅ Mapping Surfamille ret terminé. {fusion['Surfamille ret'].notna().sum()} valeurs trouvées sur {len(fusion)} lignes.")
    except Exception as e:
        print(f"[ERROR] ❌ Erreur mapping Surfamille ret : {e}")
        traceback.print_exc()
        if "Surfamille ret" not in fusion.columns:
            fusion["Surfamille ret"] = None

    # ---------------------------- ENSEIGNE RET ----------------------------
    try:
        print(f"[AUDIT] Shape de table_df juste avant mapping Enseigne ret : {table_df.shape if table_df is not None else 'table_df=None'}")
        print("[INFO] 🔍 Nouveau mapping Enseigne ret (V2 puis fallback)...")
        fusion["ENSEIGNE"] = fusion["ENSEIGNE"].fillna("").astype(str).str.strip().str.upper()
        fusion["CUSTOMER NAME"] = fusion["CUSTOMER NAME"].fillna("").astype(str).str.strip().str.upper()
        print(f"[DEBUG] Type ENSEIGNE fusion : {fusion['ENSEIGNE'].apply(type).value_counts()}")
        print(f"[DEBUG] Exemples ENSEIGNE fusion : {fusion['ENSEIGNE'].unique()[:10]}")
        if table_df is not None:
            # --- LOG: Stats sur les colonnes de mapping V2 ---
            if "ENSEIGNE V2" in table_df.columns and "Enseigne ret V2" in table_df.columns:
                table_df["ENSEIGNE V2"] = table_df["ENSEIGNE V2"].astype(str).str.strip().str.upper()
                table_df["Enseigne ret V2"] = table_df["Enseigne ret V2"].astype(str).str.strip().str.upper()
                print(f"[DEBUG] Type ENSEIGNE V2 table : {table_df['ENSEIGNE V2'].apply(type).value_counts()}")
                print(f"[DEBUG] Exemples ENSEIGNE V2 table : {table_df['ENSEIGNE V2'].unique()[:10]}")
                print(f"[LOG] Nb valeurs distinctes ENSEIGNE V2 : {table_df['ENSEIGNE V2'].nunique()}")
                print(f"[LOG] Exemples ENSEIGNE V2 : {table_df['ENSEIGNE V2'].dropna().unique()[:5]}")
                print(f"[LOG] Nb valeurs distinctes Enseigne ret V2 : {table_df['Enseigne ret V2'].nunique()}")
                print(f"[LOG] Exemples Enseigne ret V2 : {table_df['Enseigne ret V2'].dropna().unique()[:5]}")
                mapping_v2 = dict(zip(table_df["ENSEIGNE V2"], table_df["Enseigne ret V2"]))
                fusion["Enseigne ret"] = fusion["ENSEIGNE"].map(mapping_v2)
                found_v2 = fusion["Enseigne ret"].notna().sum()
                print(f"[LOG] Après mapping V2 : {found_v2} correspondances trouvées sur {len(fusion)} lignes.")
                if found_v2 > 0:
                    print(f"[LOG] Exemples de valeurs enrichies (V2) : {fusion.loc[fusion['Enseigne ret'].notna(), ['ENSEIGNE','Enseigne ret']].head(5).to_dict(orient='records')}")
                not_found_v2 = fusion.loc[fusion["Enseigne ret"].isna(), "ENSEIGNE"].unique()[:10]
                print(f"[LOG] Exemples de ENSEIGNE non trouvées en V2 : {not_found_v2}")
            else:
                fusion["Enseigne ret"] = None
            # --- Fallback ancien mapping ---
            mask_vide = fusion["Enseigne ret"].isna()
            if mask_vide.any() and "CONCAT NAME" in table_df.columns and "Enseigne ret" in table_df.columns:
                table_df["CONCAT NAME"] = table_df["CONCAT NAME"].astype(str).str.strip().str.upper()
                table_df["Enseigne ret"] = table_df["Enseigne ret"].astype(str).str.strip().str.upper()
                print(f"[DEBUG] Type CONCAT NAME table (fallback) : {table_df['CONCAT NAME'].apply(type).value_counts()}")
                print(f"[DEBUG] Exemples CONCAT NAME table (fallback) : {table_df['CONCAT NAME'].unique()[:10]}")
                fusion["concat_key"] = fusion["ENSEIGNE"] + fusion["CUSTOMER NAME"]
                print(f"[DEBUG] Exemples concat_key fusion : {fusion['concat_key'].unique()[:10]}")
                mapping_old = dict(zip(table_df["CONCAT NAME"], table_df["Enseigne ret"]))
                fusion.loc[mask_vide, "Enseigne ret"] = fusion.loc[mask_vide, "concat_key"].map(mapping_old)
                found_fallback = fusion["Enseigne ret"].notna().sum() - found_v2
                print(f"[LOG] Après fallback : {found_fallback} correspondances trouvées en plus.")
                if found_fallback > 0:
                    print(f"[LOG] Exemples de valeurs enrichies (fallback) : {fusion.loc[mask_vide & fusion['Enseigne ret'].notna(), ['ENSEIGNE','CUSTOMER NAME','Enseigne ret']].head(5).to_dict(orient='records')}")
                not_found_final = fusion.loc[fusion["Enseigne ret"].isna(), "ENSEIGNE"].unique()[:10]
                print(f"[LOG] Exemples de ENSEIGNE toujours non trouvées : {not_found_final}")
                fusion.drop(columns=["concat_key"], inplace=True, errors="ignore")
            # Si toujours rien, laisser vide
            mask_vide = fusion["Enseigne ret"].isna()
            if mask_vide.any():
                print(f"[LOG] {mask_vide.sum()} lignes sans aucune correspondance pour Enseigne ret.")
                fusion.loc[mask_vide, "Enseigne ret"] = None
        else:
            fusion["Enseigne ret"] = None
        print(f"[INFO] ✅ Mapping Enseigne ret terminé. {fusion['Enseigne ret'].notna().sum()} valeurs trouvées sur {len(fusion)} lignes.")
    except Exception as e:
        print(f"[ERROR] ❌ Erreur mapping Enseigne ret : {e}")
        traceback.print_exc()
        if "Enseigne ret" not in fusion.columns:
            fusion["Enseigne ret"] = None

    # 🔍 Extraire les dates uniques de la colonne "MONTH"
    if "MONTH" in fusion.columns:
        try:
            fusion["MONTH"] = pd.to_datetime(fusion["MONTH"], errors="coerce")
            dates_disponibles = sorted(fusion["MONTH"].dropna().dt.strftime("%Y-%m-%d").unique())
        except Exception as e:
            print(f"[ERROR] Impossible de convertir les dates : {e}")
            dates_disponibles = []
    else:
        print("[WARN] ❌ Aucune colonne 'MONTH' trouvée.")
        dates_disponibles = []

    # 📋 Afficher les dates disponibles pour que l'utilisateur les choisisse
    if dates_disponibles:
        print(f"\n🗓️ Dates détectées dans les fichiers :\n" + "\n".join(f"  • {d}" for d in dates_disponibles))
        
        if args.mois_selectionnes:
            mois_choisis = args.mois_selectionnes.split(",")
            print(f"\n✅ Mois choisis via l'interface : {mois_choisis}")
            fusion = fusion[fusion["MONTH"].dt.to_period("M").astype(str).isin(mois_choisis)]
        else:
            if os.environ.get("FROM_GUI") == "1":
                print("[ERROR] ❌ Aucun mois sélectionné et interaction impossible (lancé depuis GUI). Merci de sélectionner les mois dans l'interface.")
                sys.exit(1)
            else:
                print("\n⏳ Entrez les dates à inclure séparées par une virgule (ex: 2025-01-01,2025-01-15) :")
                user_input = input(">>> ").strip()
                dates_choisies = [d.strip() for d in user_input.split(",") if d.strip() in dates_disponibles]
                print(f"\n✅ Dates retenues : {dates_choisies}\n")
                fusion = fusion[fusion["MONTH"].dt.strftime("%Y-%m-%d").isin(dates_choisies)]

    else:
        print("[WARN] ❌ Aucune date valide détectée, aucun filtre appliqué.")


    fusion["CURRENCY"] = fusion["CURRENCY"].str.strip().str.upper()
    fusion["Taux €"] = fusion["CURRENCY"].map(rates)

    # Correction du calcul de "C.A en €" pour prendre en compte les devises avec taux > 1
    # NOUVELLE FORMULE UNIVERSELLE : Montant en EUR = Montant en DEVISE / Taux (EUR/DEV)
    fusion["C.A en €"] = fusion.apply(
        lambda row: row["TURNOVER"] / row["Taux €"]
        if pd.notnull(row.get("TURNOVER")) and pd.notnull(row.get("Taux €")) and row["Taux €"] != 0
        else None,
        axis=1
    )

    # Conversion de COGS et VARIABLE COSTS en euros (remplacement direct)
    fusion["COGS"] = fusion.apply(
        lambda row: row["COGS"] / row["Taux €"]
        if pd.notnull(row.get("COGS")) and pd.notnull(row.get("Taux €")) and row["Taux €"] != 0
        else None,
        axis=1
    )
    fusion["VARIABLE COSTS"] = fusion.apply(
        lambda row: row["VARIABLE COSTS"] / row["Taux €"]
        if pd.notnull(row.get("VARIABLE COSTS")) and pd.notnull(row.get("Taux €")) and row["Taux €"] != 0
        else None,
        axis=1
    )

    # ➕ Calcul des marges avec les coûts déjà convertis en euros
    fusion["VAR Margin"] = fusion.apply(
        lambda row: row["C.A en €"] - (row["VARIABLE COSTS"] * row["QUANTITY"])
        if pd.notnull(row.get("C.A en €")) and pd.notnull(row.get("VARIABLE COSTS")) and pd.notnull(row.get("QUANTITY"))
        else None,
        axis=1
    )

    fusion["Margin"] = fusion.apply(
        lambda row: row["C.A en €"] - (row["COGS"] * row["QUANTITY"])
        if pd.notnull(row.get("C.A en €")) and pd.notnull(row.get("COGS")) and pd.notnull(row.get("QUANTITY"))
        else None,
        axis=1
    )



    dev_non_gérées = devises_detectées - rates.keys()

    print(f"[INFO] 🏦 Devises détectées dans les fichiers : {sorted(devises_detectées)}", flush=True)
    print(f"[INFO] ✅ Taux disponibles ECB : {sorted(rates.keys())}", flush=True)

    if dev_non_gérées:
        print(f"[WARN] ⚠ Les devises suivantes n'ont pas de taux ECB : {sorted(dev_non_gérées)}", flush=True)
    else:
        print(f"[INFO] 🎉 Tous les taux de devises sont disponibles 🎯", flush=True)


    ORDER = [
    "MONTH", "SIAMP UNIT", "SALE TYPE", "TYPE OF CANAL", "ENSEIGNE", "CUSTOMER NAME",
    "COMMERCIAL AREA", "SUR FAMILLE", "FAMILLE", "REFERENCE", "PRODUCT NAME",
    "QUANTITY", "TURNOVER", "CURRENCY", "COUNTRY", "C.A en €",
    "VARIABLE COSTS", "COGS", "VAR Margin", "Margin",
    "Enseigne ret", "Surfamille ret", "NOMFICHIER", "FEUILLE", "SOURCE", "Taux €"
]


    if fusion.empty:
        print("[ERROR] ❌ Aucune donnée après le filtrage, arrêt du script.", flush=True)
        sys.exit(1)

    fusion = fusion[[c for c in ORDER if c in fusion.columns]
                    + [c for c in fusion.columns if c not in ORDER]]
    
    before = fusion.shape[0]

    # ➤ Vérification de la limite Excel (1 048 576 lignes maximum)
    if fusion.shape[0] > 1048576:
        print(f"[ERROR] ❌ Le fichier final contient {fusion.shape[0]} lignes, ce qui dépasse la limite Excel de 1 048 576 lignes.", flush=True)
        print(f"[SUGGESTION] Solutions possibles :", flush=True)
        print(f"   1. Exclure le fichier TUR 05.xlsx qui contient 1M+ lignes vides", flush=True)
        print(f"   2. Diviser le traitement en plusieurs fichiers", flush=True)
        print(f"   3. Appliquer des filtres plus stricts sur les données", flush=True)
        sys.exit(1)
    
    # ➤ Nettoyage global des chaînes : suppression espaces, mise en majuscule, suppression caractères invisibles
    def nettoyer_str(s):
        if pd.isna(s):
            return None
        if isinstance(s, str):
            s = s.strip().upper()
            s = re.sub(r'[^\x20-\x7E\u00A0-\uFFFF]', '', s)  # supprime caractères invisibles
            return s
        return s

    # Appliquer le nettoyage sur toutes les colonnes objet
    for col in fusion.select_dtypes(include="object").columns:
        fusion[col] = fusion[col].apply(nettoyer_str)

    print(f"[AUDIT] Shape juste avant export Excel : {fusion.shape}", flush=True)

    # ➤ Nettoyage global des chaînes : suppression espaces, mise en majuscule, suppression caractères invisibles
    def nettoyer_str(s):
        if pd.isna(s):
            return None
        if isinstance(s, str):
            s = s.strip().upper()
            s = re.sub(r'[^\x20-\x7E\u00A0-\uFFFF]', '', s)  # supprime caractères invisibles
            return s
        return s

    # Appliquer le nettoyage sur toutes les colonnes objet
    for col in fusion.select_dtypes(include="object").columns:
        fusion[col] = fusion[col].apply(nettoyer_str)

    # Tentative d'export Excel avec gestion d'erreur robuste
    max_attempts = 3
    for attempt in range(max_attempts):
        try:
            print(f"[INFO] Tentative d'export Excel #{attempt + 1}/{max_attempts}...", flush=True)
            
            # Vérifier si le fichier existe et est ouvert
            if os.path.exists(out):
                try:
                    # Essayer d'ouvrir le fichier en mode écriture pour vérifier les permissions
                    with open(out, 'a') as test_file:
                        pass
                except PermissionError:
                    print(f"[ERROR] Le fichier {out} est ouvert dans Excel ou un autre programme.", flush=True)
                    print("[SUGGESTION] Fermez le fichier Excel et relancez le script.", flush=True)
                    if attempt < max_attempts - 1:
                        print(f"[INFO] Nouvelle tentative dans 2 secondes...", flush=True)
                        sleep(2)
                        continue
                    else:
                        raise PermissionError(f"Impossible d'écrire dans {out} après {max_attempts} tentatives")
            
            # Créer le répertoire de sortie s'il n'existe pas
            output_dir = os.path.dirname(out)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir, exist_ok=True)
                print(f"[INFO] Répertoire créé : {output_dir}", flush=True)
            
            # Export Excel
            fusion.to_excel(out, index=False)
            print(f"[AUDIT] Shape après export Excel : {fusion.shape}", flush=True)
            print(f"[SUCCESS] ✅ Fichier Excel créé avec succès : {out}", flush=True)
            break
            
        except PermissionError as e:
            print(f"[ERROR] Tentative #{attempt + 1} échouée - Permission refusée : {e}", flush=True)
            if attempt < max_attempts - 1:
                print("[SUGGESTION] Vérifiez que :", flush=True)
                print("  1. Le fichier n'est pas ouvert dans Excel", flush=True)
                print("  2. Vous avez les droits d'écriture dans le dossier", flush=True)
                print("  3. Le fichier n'est pas en lecture seule", flush=True)
                print(f"[INFO] Nouvelle tentative dans 3 secondes...", flush=True)
                sleep(3)
            else:
                print("[FATAL ERROR] Impossible de créer le fichier Excel après toutes les tentatives.", flush=True)
                print("[SOLUTIONS POSSIBLES] :", flush=True)
                print("  1. Fermez Excel et tous les programmes qui pourraient utiliser le fichier", flush=True)
                print("  2. Choisissez un autre nom de fichier ou un autre emplacement", flush=True)
                print("  3. Vérifiez les permissions du dossier de destination", flush=True)
                print("  4. Exécutez le script en tant qu'administrateur si nécessaire", flush=True)
                raise
        except Exception as e:
            print(f"[ERROR] Erreur inattendue lors de l'export Excel : {e}", flush=True)
            if attempt < max_attempts - 1:
                print(f"[INFO] Nouvelle tentative dans 2 secondes...", flush=True)
                sleep(2)
            else:
                raise

    # mise en forme Excel
    print("[DEBUG] 🟡 Début de la mise en forme Excel...", flush=True)
    try:
        wb = load_workbook(out)
        ws = wb.active

        print(f"[DEBUG] 📊 Workbook chargé : {out}", flush=True)
        print(f"[DEBUG] Nombre de lignes : {ws.max_row}, Nombre de colonnes : {ws.max_column}", flush=True)

        if ws.max_row > 1 and ws.max_column > 0:
            last_col_letter = get_column_letter(ws.max_column)
            last_row = ws.max_row
            table_range = f"A1:{last_col_letter}{last_row}"
            print(f"[DEBUG] 🖋️ Définition de la table FusionTable sur la plage : {table_range}", flush=True)

            table = Table(displayName="FusionTable", ref=table_range)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            
            # ─── Videz d'abord toute table existante ───────────────────────
            ws._tables.clear()

            # ─── Ajout de la nouvelle table ───────────────────────────────
            ws.add_table(table)
            print("[DEBUG] ✅ Nouvelle table 'FusionTable' ajoutée avec succès", flush=True)


            # ➕ Formatage des colonnes €
            EURO_COLUMNS = {"C.A en €", "VAR Margin", "Margin"}
            print("[DEBUG] 🎯 Formatage des colonnes €...", flush=True)
            for col_idx in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col_idx).value
                if header in EURO_COLUMNS:
                    for row_idx in range(2, last_row + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.number_format = u"#,##0.00\u00a0€"
            print("[DEBUG] ✅ Formatage des colonnes € terminé", flush=True)
        else:
            print("[WARN] ⚠️ Impossible d'ajouter la table : pas assez de données (0 colonne ou 1 ligne).", flush=True)

        wb.save(out)
        print(f"[SUCCESS] ✅ Mise en forme Excel terminée avec succès", flush=True)
        
        if fichiers_ignores:
            print(f"\n⚠️ Fusion partielle : certains fichiers n'ont pas été traités à cause de colonnes non conformes :", flush=True)
            for f in fichiers_ignores:
                print(f"   - {f['fichier']}", flush=True)
                print(f"     Motif : {f['motif']}", flush=True)
                if f['colonnes_manquantes']:
                    print(f"     Colonnes manquantes : {f['colonnes_manquantes']}", flush=True)
                if f['colonnes_sup']:
                    print(f"     Colonnes supplémentaires : {f['colonnes_sup']}", flush=True)
            print(f"\n⚠️ Fusion terminée avec des fichiers ignorés. Voir détails ci-dessus.\n", flush=True)
        else:
            print(f"\n✅ Fusion terminée – fichier créé : {out}\n", flush=True)

    except Exception as e:
        print(f"[ERROR] ❌ Une erreur s'est produite pendant la mise en forme Excel : {e}", flush=True)
        print("[INFO] Le fichier Excel a été créé mais sans mise en forme.", flush=True)
        # Ne pas arrêter le script ici, le fichier a été créé avec succès

def validate_strict_columns(df, filename, formats, return_details=False):
    """
    Valide la présence des colonnes requises sans être strict sur leur ordre.
    Affiche un log détaillé en cas de problème.
    Si return_details=True, retourne (is_valid, motif, colonnes_manquantes, colonnes_sup)
    """
    def norm(col):
        return col.strip().replace(" ", "").replace("_", "").upper()
    
    cols = [norm(c) for c in df.columns]
    
    # Pour chaque format, vérifier si toutes les colonnes requises sont présentes
    for fmt in formats:
        fmt_norm = [norm(c) for c in fmt]
        missing = [c for c in fmt if norm(c) not in cols]
        
        # Si aucune colonne ne manque pour ce format, le fichier est valide
        if not missing:
            if return_details:
                return True, '', [], []
            return True
    
    # Si aucun format ne correspond complètement, on détaille
    motif = "Colonnes manquantes"
    cols_manquantes = []
    for fmt in formats:
        missing = [c for c in fmt if norm(c) not in cols]
        if missing:
            cols_manquantes.extend(missing)
    
    # Éliminer les doublons dans les colonnes manquantes
    cols_manquantes = list(set(cols_manquantes))
    
    cols_sup = [c for c in df.columns if norm(c) not in [norm(x) for f in formats for x in f]]
    
    if return_details:
        return False, motif, cols_manquantes, cols_sup
    
    # Affichage classique (non utilisé ici)
    print(f"[ERREUR COLONNES] {filename}")
    print(f"  Colonnes trouvées : {df.columns.tolist()}")
    if cols_manquantes:
        print(f"  Colonnes manquantes : {cols_manquantes}")
    if cols_sup:
        print(f"  Colonnes supplémentaires : {cols_sup}")
    
    return False

# --------------------------------------------------
# Lancement sécurisé du script avec capture des erreurs
# --------------------------------------------------
if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"[FATAL ERROR] ❌ Le script a planté avec l'exception : {e}", flush=True)
        import traceback
        traceback.print_exc()
        sys.exit(1)