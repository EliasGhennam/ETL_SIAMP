#executer avec : pyinstaller --noconsole --onefile --icon=siamp_icon.ico --version-file=version.txt ETL_SIAMP.py

# -*- coding: utf-8 -*-
from gooey import Gooey, GooeyParser
import pandas as pd
import os
import re
from time import sleep
import warnings
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import sys
import io

if sys.stdout and hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

@Gooey(
    program_name="Fusion Excel - SIAMP",
    language='french',
    default_size=(720, 640),
    navigation='TABBED',
    required_cols=1,
    optional_cols=2,
    show_success_modal=True,
    show_failure_modal=True,
    clear_before_run=True,
    show_progress_bar=True,
    progress_regex=r"^PROGRESS: (\d+)%$",
    use_legacy_titles=True
)
def main():
    parser = GooeyParser(description="Fusionnez plusieurs fichiers Excel contenant des feuilles nommées Turnover")

    parser.add_argument(
        "fichiers",
        metavar="Fichiers Excel à fusionner",
        widget="MultiFileChooser",
        help="Sélectionnez les fichiers Excel à traiter (.xlsx uniquement)",
        nargs="+",
        gooey_options={
            'wildcard': "Fichiers Excel (*.xlsx)|*.xlsx"
        }
    )

    parser.add_argument(
        "chemin_sortie",
        metavar="Emplacement du fichier de sortie",
        widget="FileSaver",
        help="Chemin complet du fichier final (ex: C:\\...\\fusion.xlsx)",
        default="fusion_finale.xlsx",
        gooey_options={
            'wildcard': "Fichiers Excel (*.xlsx)|*.xlsx"
        }
    )

    args = parser.parse_args()

    fichiers = [
        f for f in args.fichiers
        if f.endswith('.xlsx') and not os.path.basename(f).startswith('~$')
    ]
    
    chemin_final = args.chemin_sortie
    dossier_sortie = os.path.dirname(chemin_final)

    if not dossier_sortie:
        dossier_sortie = "."
        chemin_final = os.path.join(dossier_sortie, chemin_final)

    os.makedirs(dossier_sortie, exist_ok=True)

    # ✅ Ajout automatique de l'extension si manquante
    if not chemin_final.lower().endswith(".xlsx"):
        chemin_final += ".xlsx"

    pattern_turnover = re.compile(r"^Turnover$|^TURNOVER$|^Turnover\s+[A-Z][a-z]{2}\s+\d{1,2}$")
    dfs = []
    total = len(fichiers)

    print("Début de la fusion des fichiers...\n")

    for i, fichier in enumerate(fichiers):
        print(f"🔍 Analyse du fichier : {os.path.basename(fichier)}", flush=True)
        try:
            xls = pd.ExcelFile(fichier, engine="openpyxl")
            feuilles = [s for s in xls.sheet_names if pattern_turnover.match(s)]

            if not feuilles:
                print(f"⚠️ Aucune feuille Turnover détectée dans {os.path.basename(fichier)}. Vérifiez son format.", flush=True)
                continue

            for feuille in feuilles:
                print(f"✅ Feuille trouvée : {feuille} ({os.path.basename(fichier)})", flush=True)
                df = xls.parse(feuille, usecols="A:O")

                if "CURRENCY" in df.columns and "Currency" not in df.columns:
                    df.rename(columns={"CURRENCY": "Currency"}, inplace=True)
                elif "Currency" in df.columns and "CURRENCY" in df.columns:
                    df["Currency"] = df["Currency"].combine_first(df["CURRENCY"])
                    df.drop(columns=["CURRENCY"], inplace=True)

                if "CUSTOMER NAME" in df.columns and "Customer" not in df.columns:
                    df.rename(columns={"CUSTOMER NAME": "Customer Name"}, inplace=True)
                elif "Customer" in df.columns and "CUSTOMER NAME" in df.columns:
                    df["Customer Name"] = df["Customer"].combine_first(df["CUSTOMER NAME"])
                    df.drop(columns=["Customer", "CUSTOMER NAME"], inplace=True)
                elif "Customer" in df.columns:
                    df.rename(columns={"Customer": "Customer Name"}, inplace=True)

                df["NomFichier"] = os.path.basename(fichier)
                df["Feuille"] = feuille
                dfs.append(df)

                print(f"[OK] Feuille ajoutée : {feuille} ({os.path.basename(fichier)})")

        except Exception as e:
            print(f"[ERREUR] Problème avec {fichier} : {e}")

        pourcentage = int(((i + 1) / total) * 100)
        if pourcentage == 100:
            print("\n⏳ Les données sont entièrement chargées. Veuillez patienter pendant la finalisation du fichier Excel (ne fermez pas l'application)...", flush=True)
        
        sleep(0.1)

    if not dfs:
        print("\nAucun fichier ou feuille valide détecté. Arrêt.")
        return

    fusion = pd.concat(dfs, ignore_index=True)
    fusion.to_excel(chemin_final, index=False)

    wb = load_workbook(chemin_final)
    ws = wb.active
    max_col = ws.max_column
    max_row = ws.max_row
    end_col_letter = get_column_letter(max_col)
    table_range = f"A1:{end_col_letter}{max_row}"

    table = Table(displayName="FusionTable", ref=table_range)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)

    wb.save(chemin_final)

    recap = "\n=== ✅ FUSION COMPLÉTÉE AVEC SUCCÈS ===\n"
    recap += f"📄 Fichier généré : {chemin_final}\n"

    recap += "\nMerci d’avoir utilisé l’outil ETL SIAMP. 🚀\n"

    print(recap, flush=True)

    

if __name__ == '__main__':
    main()
