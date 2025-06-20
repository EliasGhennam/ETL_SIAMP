#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
ETL_SIAMP_GUI.py – Interface PyQt6 améliorée
----------------------------------------------
• Sélecteur de date + chargement historique des taux.
• Glisser‑déposer de fichiers Excel + ajout/retrait.
• Console en temps réel + barre de progression.
• Exécute le script core `ETL_SIAMP.py` via subprocess.
"""
from __future__ import annotations
import os
import sys
import requests

# Ajout du système de logging pour le débogage
import logging
import tempfile
import datetime

# Configurer le logging pour capturer toutes les erreurs
log_file = os.path.join(tempfile.gettempdir(), f"etl_siamp_log_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.txt")
logging.basicConfig(
    filename=log_file,
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)

# Rediriger les exceptions non capturées vers le log
def handle_exception(exc_type, exc_value, exc_traceback):
    logging.error("Exception non capturée", exc_info=(exc_type, exc_value, exc_traceback))
    # Afficher dans la console si disponible
    print(f"ERREUR CRITIQUE: {exc_value}")

sys.excepthook = handle_exception

# Log initial avec informations système
logging.info("=== Application ETL_SIAMP_GUI démarrée ===")
logging.info(f"Python version: {sys.version}")
logging.info(f"Python executable: {sys.executable}")
logging.info(f"Current working directory: {os.getcwd()}")
logging.info(f"System platform: {sys.platform}")
logging.info(f"Environment variables:")
for var in ['PATH', 'PYTHONPATH', 'TEMP', 'TMP']:
    logging.info(f"  {var}: {os.environ.get(var, 'Non défini')}")

# Vérifier si on est dans un environnement PyInstaller
if hasattr(sys, '_MEIPASS'):
    logging.info(f"PyInstaller directory: {sys._MEIPASS}")
    logging.info(f"PyInstaller files:")
    for root, dirs, files in os.walk(sys._MEIPASS):
        for file in files:
            logging.info(f"  {os.path.join(root, file)}")
else:
    logging.info("Pas dans un environnement PyInstaller")

def resource_path(relative_path):
    if hasattr(sys, "_MEIPASS"):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)


import re
import pandas as pd
import configparser
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import subprocess
import shutil
import calendar
from typing import List
import xml.etree.ElementTree as ET
from datetime import datetime
import requests
from PyQt6.QtCore   import Qt, QThread, pyqtSignal, QDate
from PyQt6.QtGui    import QIcon, QAction, QKeySequence, QPainter, QFont, QColor
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QLineEdit, QPushButton, QFileDialog, QMessageBox, QListWidget, QComboBox,
    QPlainTextEdit, QProgressBar, QDateEdit, QInputDialog, QFrame, QScrollArea
)

# Forcer le style sombre indépendamment du thème système
os.environ['QT_QPA_PLATFORM'] = 'windows:darkmode=2'

SCRIPT_CORE = "ETL_SIAMP.py"
ICON_PATH        = resource_path("mydata/siamp_icon.ico")
CONFIG_FILE      = resource_path("mydata/siamp_api_key.cfg")
CONFIG_REF_FILE  = resource_path("mydata/ref_files.cfg")

# Définir un mapping de colonnes standard
COLUMN_MAPPING = {
    # Variations possibles -> Nom standardisé
    "MONTH": ["MONTH", "DATE", "PERIODE"],
    "CUSTOMER NAME": ["CUSTOMER NAME", "CUSTOMER", "CLIENT", "NOM CLIENT"],
    "REFERENCE": ["REFERENCE", "REF", "REFERENCE PRODUIT"],
    "TURNOVER": ["TURNOVER", "CA", "CHIFFRE D'AFFAIRE", "SALES"],
    "QUANTITY": ["QUANTITY", "QTY", "QUANTITE"],
    "CURRENCY": ["CURRENCY", "DEVISE", "MONNAIE"]
}

# ---------------------------------------------------------------- worker QThread
class Worker(QThread):
    log      = pyqtSignal(str)
    progress = pyqtSignal(int)
    done     = pyqtSignal(bool)

    def __init__(self, cmd: list[str], env: dict[str,str]):
        super().__init__()
        self.cmd = cmd
        self.env = env
        logging.info(f"Worker initialisé avec cmd: {cmd}")

    def run(self):
        log_path = os.path.join(tempfile.gettempdir(), "etl_siamp_subprocess.log")
        logging.info(f"Log du subprocess: {log_path}")
        
        try:
            with open(log_path, "w", encoding="utf-8", errors="replace") as err_file:
                logging.info("Début de l'exécution du subprocess")
                
                # Ajouter ces lignes pour masquer la console
                startupinfo = None
                if sys.platform == "win32":
                    startupinfo = subprocess.STARTUPINFO()
                    startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
                
                logging.info("Création du processus")
                proc = subprocess.Popen(
                    self.cmd,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE,  # Capturer stderr aussi
                    text=True,
                    env=self.env,
                    encoding='utf-8',
                    errors='replace',
                    startupinfo=startupinfo
                )
                
                # Traiter stdout
                for line in proc.stdout:
                    line = line.rstrip()
                    self.log.emit(line)
                    logging.debug(f"STDOUT: {line}")
                    if line.startswith("PROGRESS:"):
                        try:
                            pct = int(line.split(":")[1].strip().strip("% "))
                            self.progress.emit(pct)
                        except ValueError:
                            pass
                
                # Capturer stderr
                stderr_output, _ = proc.communicate()
                if stderr_output:
                    logging.error(f"STDERR: {stderr_output}")
                    self.log.emit(f"[ERREUR] {stderr_output}")
                    err_file.write(f"STDERR: {stderr_output}\n")
                
                return_code = proc.wait()
                logging.info(f"Processus terminé avec code: {return_code}")
                
                if return_code != 0:
                    logging.error(f"Échec du processus avec code: {return_code}")
                    self.log.emit(f"[ERREUR] Le processus a échoué avec le code {return_code}")
                    
                self.done.emit(return_code == 0)
                
        except Exception as e:
            logging.exception("Exception dans le Worker")
            self.log.emit(f"[ERREUR CRITIQUE] {e}")
            import traceback
            tb_str = traceback.format_exc()
            logging.error(f"Traceback: {tb_str}")
            self.log.emit(f"[TRACEBACK] {tb_str}")
            self.done.emit(False)


# ---------------------------------------------------------------- DropListWidget
class DropListWidget(QListWidget):
    """Zone de liste acceptant le glisser‑déposer de fichiers .xlsx"""

    def __init__(self, on_click_callback=None):
        super().__init__()
        self.setAcceptDrops(True)
        self.setSelectionMode(self.SelectionMode.ExtendedSelection)
        self.setMinimumHeight(150)
        self.on_click_callback = on_click_callback  # fonction à appeler au clic

    def paintEvent(self, event):
        super().paintEvent(event)
        if self.count() == 0:
            painter = QPainter(self.viewport())
            painter.setPen(QColor("#777"))
            font = QFont("Segoe UI", 10, QFont.Weight.Normal)
            font.setItalic(True)
            painter.setFont(font)
            text = "Glissez vos fichiers Excel ici ou cliquez pour les sélectionner"
            painter.drawText(self.rect(), Qt.AlignmentFlag.AlignCenter, text)

    def mousePressEvent(self, event):
        if self.count() == 0 and self.on_click_callback:
            self.on_click_callback()  # déclenche la fonction ajout fichiers
        super().mousePressEvent(event)

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dragMoveEvent(self, event):
        event.acceptProposedAction()

    def dropEvent(self, event):
        for url in event.mimeData().urls():
            f = url.toLocalFile()
            if f.lower().endswith(".xlsx") and f not in self.files():
                self.addItem(f)
        event.acceptProposedAction()

    def files(self) -> List[str]:
        return [self.item(i).text() for i in range(self.count())]

class ColumnStatusBar(QFrame):
    """Bandeau affichant le statut de détection des colonnes."""
    
    def __init__(self, expected_columns, parent=None):
        super().__init__(parent)
        self.expected_columns = expected_columns
        self.column_labels = {}
        self.setFrameStyle(QFrame.Shape.Panel | QFrame.Shadow.Sunken)
        
        # Layout horizontal avec scrolling si nécessaire
        layout = QHBoxLayout(self)
        layout.setContentsMargins(5, 5, 5, 5)
        layout.setSpacing(5)
        
        # Création des labels pour chaque colonne
        for col in expected_columns:
            label = QLabel(col)
            label.setFixedHeight(20)  # Plus petit
            label.setMinimumWidth(90)  # Largeur minimale pour le label
            label.setMaximumWidth(110)  # Largeur max, à ajuster selon ton UI
            label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            label.setWordWrap(False)
            label.setStyleSheet("""
                background-color: #3D444C;
                color: #AAAAAA; 
                border-radius: 4px;
                padding: 1px 6px;
                margin: 1px;
                font-size: 9pt;
            """)
            self.column_labels[col] = label
            layout.addWidget(label)
        
        # Empêcher le redimensionnement vertical
        self.setMaximumHeight(40)
        
        # Ajouter un stretch à la fin pour que les labels restent alignés à gauche
        layout.addStretch()

    @staticmethod
    def create_scrollable(expected_columns, parent=None):
        """Crée un widget avec barre de défilement horizontale contenant le bandeau de colonnes."""
        scroll_area = QScrollArea(parent)
        scroll_area.setWidgetResizable(True)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        scroll_area.setFrameShape(QFrame.Shape.NoFrame)  # Supprime le cadre du QScrollArea
        
        # Créer le bandeau et l'ajouter au scroll area
        column_bar = ColumnStatusBar(expected_columns)
        scroll_area.setWidget(column_bar)
        
        # Limiter la hauteur du scroll area
        scroll_area.setMaximumHeight(40)
        
        # Créer un widget conteneur pour ajouter une marge
        container = QWidget()
        container_layout = QVBoxLayout(container)
        container_layout.setContentsMargins(0, 0, 0, 0)
        container_layout.setSpacing(0)
        container_layout.addWidget(scroll_area)
        
        return container, column_bar

    def update_status_interactive(self, presence, all_files):
        for col, label in self.column_labels.items():
            files_with = presence.get(col, set())
            missing = [os.path.basename(f) for f in all_files if f not in files_with]
            if len(files_with) == len(all_files) and all_files:
                # Vert : présent partout
                label.setStyleSheet("background-color: #297F4F; color: white; border-radius: 4px; font-weight: bold; font-size:9pt;")
                label.setToolTip("Présent dans tous les fichiers")
                label.mousePressEvent = None
            elif len(files_with) == 0:
                # Rouge : absent partout
                label.setStyleSheet("background-color: #B22222; color: white; border-radius: 4px; font-weight: bold; font-size:9pt;")
                if missing:
                    label.setToolTip("Absent de tous les fichiers :\n" + "\n".join(missing))
                else:
                    label.setToolTip("Absent de tous les fichiers")
                label.mousePressEvent = None
            else:
                # Orange : partiel
                label.setStyleSheet("background-color: #FFA500; color: black; border-radius: 4px; font-weight: bold; font-size:9pt;")
                if missing:
                    label.setToolTip("Manquant dans :\n" + "\n".join(missing))
                else:
                    label.setToolTip("Présent partiellement")
                label.mousePressEvent = None  # On n'a plus besoin du clic


# ---------------------------------------------------------------- MainWindow
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("ETL SIAMP — Fusion Excel")
        self.setWindowIcon(QIcon(ICON_PATH))
        self.resize(760, 640)
        self._build_tabs()
        self._apply_style()
        self.check_for_update()

    def _detect_months(self):
        from collections import defaultdict
        from PyQt6.QtWidgets import QDialog, QTreeWidget, QTreeWidgetItem, QVBoxLayout, QPushButton

        # Structure hiérarchique pour stocker année > mois > jour
        dates_hierarchie = defaultdict(lambda: defaultdict(set))
        files = self.lst_files.files()

        if not files:
            QMessageBox.warning(self, "Erreur", "Ajoutez au moins un fichier Excel.")
            return

        # ➤ Détection des dates dans les fichiers
        for path in files:
            try:
                xls = pd.ExcelFile(path, engine="openpyxl")
                for sh in xls.sheet_names:
                    df = xls.parse(sh, usecols="A:Q")
                    df.columns = [c.strip().upper() for c in df.columns]
                    if "MONTH" in df.columns:
                        # Conversion robuste en dates
                        dates = pd.to_datetime(df["MONTH"], errors="coerce")
                        # Pour chaque date valide, extraire année/mois/jour
                        for date in dates.dropna():
                            dates_hierarchie[date.year][date.month].add(date.day)
            except Exception as e:
                self.txt_log.appendPlainText(f"[WARN] ⚠ Fichier ignoré : {path} – {e}")

        if not dates_hierarchie:
            QMessageBox.information(self, "Info", "Aucune date détectée dans les fichiers.")
            return

        # ➤ Création de la boîte de dialogue
        dialog = QDialog(self)
        dialog.setWindowTitle("Sélectionnez les mois à traiter")
        layout = QVBoxLayout(dialog)
        tree = QTreeWidget()
        tree.setHeaderLabel("Mois détectés")
        tree.setColumnCount(1)
        tree.setSelectionMode(QTreeWidget.SelectionMode.MultiSelection)
        tree.setExpandsOnDoubleClick(True)

        # ➤ Construction de l'arborescence depuis notre hiérarchie
        for annee in sorted(dates_hierarchie.keys()):
            parent = QTreeWidgetItem([str(annee)])
            parent.setFlags(parent.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            parent.setCheckState(0, Qt.CheckState.Checked)
            
            for mois_num in sorted(dates_hierarchie[annee].keys()):
                mois_nom = calendar.month_name[mois_num].capitalize()
                child = QTreeWidgetItem([mois_nom])
                child.setFlags(child.flags() | Qt.ItemFlag.ItemIsUserCheckable)
                child.setCheckState(0, Qt.CheckState.Checked)
                # Stocke le numéro du mois formaté avec leading zero
                child.setData(0, Qt.ItemDataRole.UserRole, f"{mois_num:02d}")
                parent.addChild(child)
            
            tree.addTopLevelItem(parent)

        layout.addWidget(tree)
        btn_ok = QPushButton("Valider")
        btn_ok.clicked.connect(dialog.accept)
        layout.addWidget(btn_ok)
        dialog.exec()

        # ➤ Extraire les dates cochées
        dates_choisies = []
        for i in range(tree.topLevelItemCount()):
            parent = tree.topLevelItem(i)
            annee = parent.text(0)
            for j in range(parent.childCount()):
                child = parent.child(j)
                if child.checkState(0) == Qt.CheckState.Checked:
                    mois = child.data(0, Qt.ItemDataRole.UserRole)
                    dates_choisies.append(f"{annee}-{mois}")
        
        self.mois_selectionnes = dates_choisies
        self.txt_log.appendPlainText(f"✅ Mois choisis : {self.mois_selectionnes}")

    def _build_tabs(self):
        from PyQt6.QtWidgets import QTabWidget

        self.tabs = QTabWidget()

        # Onglet 1 : Traitement mensuel (ce que tu avais déjà)
        self.page_traitement = QWidget()
        self.tabs.addTab(self.page_traitement, "Traitement Mensuel")
        self._build_traitement_ui(self.page_traitement)  # ⚠️ on utilise maintenant page_traitement ici

        # Onglet 2 : Fusion historique (nouvel onglet)
        self.page_historique = QWidget()
        self.tabs.addTab(self.page_historique, "Fusion Historique")
        self._build_historique_ui(self.page_historique)  # ⚠️ méthode à créer juste après
        self.setCentralWidget(self.tabs)

        # Onglet 3 : Paramètres (NOUVEAU)
        self.page_parametres = QWidget()
        self.tabs.addTab(self.page_parametres, "Paramètres / Références")
        self._build_parametres_ui(self.page_parametres)  # 👈 à créer juste après

    def _build_historique_ui(self, parent_widget):
        layout = QVBoxLayout(parent_widget)

        # Fichiers historiques
        layout.addWidget(QLabel("Fichiers historiques à fusionner :"))
        self.lst_historique_files = DropListWidget(on_click_callback=self._add_historique_files)
        layout.addWidget(self.lst_historique_files)

        btn_bar = QHBoxLayout()
        btn_add = QPushButton("Ajouter…")
        btn_add.clicked.connect(self._add_historique_files)
        btn_rem = QPushButton("Retirer sélection")
        btn_rem.clicked.connect(self._remove_historique_files)
        btn_bar.addWidget(btn_add)
        btn_bar.addWidget(btn_rem)
        btn_bar.addStretch()
        layout.addLayout(btn_bar)
        self.lst_historique_files.setAlternatingRowColors(True)
        
        # Bandeau des colonnes historiques avec défilement horizontal - déplacé juste après la zone d'insertion de fichiers
        layout.addWidget(QLabel("<b>Colonnes attendues (historique) :</b>"))
        expected_histo_columns = [
            "MONTH", "SIAMP UNIT", "SALE TYPE", "TYPE OF CANAL", "ENSEIGNE", "CUSTOMER NAME", "COMMERCIAL AREA",
            "SUR FAMILLE", "FAMILLE", "REFERENCE", "PRODUCT NAME", "QUANTITY", "TURNOVER", "CURRENCY",
            "COUNTRY", "C.A en €", "VARIABLE COSTS", "COGS", "VAR Margin", "Margin",
            "NOMFICHIER", "FEUILLE", "Enseigne ret", "Surfamille ret", "SOURCE", "Taux €"
        ]
        scroll_area_histo, self.histo_column_status_bar = ColumnStatusBar.create_scrollable(expected_histo_columns)
        layout.addWidget(scroll_area_histo)

        btn_check_histo = QPushButton("Vérifier le contenu")
        btn_check_histo.clicked.connect(self._check_histo_columns_in_files)
        layout.addWidget(btn_check_histo)

        # Chemin de sortie
        row_out = QHBoxLayout()
        row_out.addWidget(QLabel("Fichier de sortie :"))
        self.txt_historique_out = QLineEdit("Historique_Consolide.xlsx")
        btn_out = QPushButton("Parcourir…")
        btn_out.clicked.connect(self._choose_historique_output)
        row_out.addWidget(self.txt_historique_out)
        row_out.addWidget(btn_out)
        layout.addLayout(row_out)

        # Barre de progression + bouton lancer
        self.pbar_historique = QProgressBar()
        self.pbar_historique.setMaximum(100)
        self.pbar_historique.setValue(0)
        layout.addWidget(self.pbar_historique)

        btn_run = QPushButton("▶ Fusionner l'historique")
        btn_run.setMinimumHeight(38)
        btn_run.clicked.connect(self._run_historique_fusion)
        layout.addWidget(btn_run)

        # Console historique
        self.txt_log_historique = QPlainTextEdit()
        self.txt_log_historique.setReadOnly(True)
        self.txt_log_historique.setMaximumBlockCount(1000)
        layout.addWidget(self.txt_log_historique, stretch=2)

    def _add_historique_files(self):
        files, _ = QFileDialog.getOpenFileNames(self, "Sélectionner fichiers historiques", "", "Excel (*.xlsx)")
        for f in files:
            if f not in self.lst_historique_files.files():
                self.lst_historique_files.addItem(f)

    def _remove_historique_files(self):
        for item in self.lst_historique_files.selectedItems():
            self.lst_historique_files.takeItem(self.lst_historique_files.row(item))

    def _choose_historique_output(self):
        path, _ = QFileDialog.getSaveFileName(self, "Fichier de sortie historique", self.txt_historique_out.text(), "Excel (*.xlsx)")
        if path:
            self.txt_historique_out.setText(path)

    def _extract_year_from_filename(self, filename):
        """Extrait l'année du nom du fichier (ex: STATS 2024.xlsx -> 2024)"""
        match = re.search(r'20\d{2}', filename)
        if match:
            return int(match.group())
        return None

    def _format_date_column(self, df, year=None):
        """Formate la colonne MONTH en gérant les différents formats de date possibles"""
        if "MONTH" not in df.columns:
            return df

        # Créer une copie de la colonne pour préserver les données originales
        df["DATE_TEMP"] = df["MONTH"].astype(str)

        # 1. D'abord essayer de parser comme date complète
        try:
            dates = pd.to_datetime(df["DATE_TEMP"], format='%d/%m/%Y', errors='coerce')
            mask_fr = dates.notna()
            if mask_fr.any():
                df.loc[mask_fr, "DATE_TEMP"] = dates[mask_fr].dt.strftime("%d/%m/%Y")
        except:
            pass

        try:
            dates = pd.to_datetime(df["DATE_TEMP"], errors='coerce')
            mask_other = dates.notna()
            if mask_other.any():
                df.loc[mask_other, "DATE_TEMP"] = dates[mask_other].dt.strftime("%d/%m/%Y")
        except:
            pass

        # 2. Pour les valeurs qui sont des nombres (incluant les décimales), convertir en date avec l'année du fichier
        if year:
            # Convertir en numérique et arrondir pour gérer les .0 ou .00
            df["DATE_TEMP_NUM"] = pd.to_numeric(df["DATE_TEMP"], errors='coerce')
            numeric_mask = df["DATE_TEMP_NUM"].notna()
            if numeric_mask.any():
                try:
                    # Arrondir et vérifier si dans la plage 1-12
                    df.loc[numeric_mask, "DATE_TEMP_NUM"] = df.loc[numeric_mask, "DATE_TEMP_NUM"].round()
                    valid_months = (df["DATE_TEMP_NUM"] >= 1) & (df["DATE_TEMP_NUM"] <= 12)
                    if valid_months.any():
                        df.loc[valid_months, "DATE_TEMP"] = pd.to_datetime(
                            df.loc[valid_months, "DATE_TEMP_NUM"].apply(
                                lambda x: f"01/{int(x):02d}/{year}"
                            ),
                            format="%d/%m/%Y"
                        ).dt.strftime("%d/%m/%Y")
                except:
                    pass
            
            df = df.drop(columns=["DATE_TEMP_NUM"])

        # Remplacer l'ancienne colonne MONTH
        df["MONTH"] = df["DATE_TEMP"]
        df = df.drop(columns=["DATE_TEMP"])
        return df

    def _run_historique_fusion(self):
        files = self.lst_historique_files.files()
        if not files:
            return QMessageBox.warning(self, "Erreur", "Ajoutez au moins un fichier Excel à fusionner.")

        out = self.txt_historique_out.text().strip()
        if not out:
            return QMessageBox.warning(self, "Erreur", "Spécifiez le fichier de sortie.")

        try:
            self.txt_log_historique.clear()
            self.pbar_historique.setValue(0)
            all_dfs = []

            total = len(files)
            for idx, path in enumerate(files, 1):
                self.txt_log_historique.appendPlainText(f"[{idx}/{total}] Lecture : {os.path.basename(path)}")
                df = pd.read_excel(path, engine="openpyxl")
                
                # Extraire l'année du nom de fichier
                year = self._extract_year_from_filename(os.path.basename(path))
                
                # Formater les dates
                df = self._format_date_column(df, year)
                
                all_dfs.append(df)
                self.pbar_historique.setValue(int((idx / total) * 100))
            
            if not all_dfs:
                self.txt_log_historique.appendPlainText("❌ Aucun fichier valide à fusionner.")
                return

            fusion = pd.concat(all_dfs, ignore_index=True)

            # Réordonner les colonnes comme dans l'ETL
            ORDER = [
                "MONTH", "SIAMP UNIT", "SALE TYPE", "TYPE OF CANAL", "ENSEIGNE", "CUSTOMER NAME",
                "COMMERCIAL AREA", "SUR FAMILLE", "FAMILLE", "REFERENCE", "PRODUCT NAME",
                "QUANTITY", "TURNOVER", "CURRENCY", "COUNTRY", "C.A en €",
                "VARIABLE COSTS", "COGS", "VAR Margin", "Margin",
                "Enseigne ret", "Surfamille ret", "NOMFICHIER", "FEUILLE", "SOURCE", "Taux €"
            ]
            fusion = fusion[[c for c in ORDER if c in fusion.columns] +
                            [c for c in fusion.columns if c not in ORDER]]
            
            # ➤ Réorganisation des colonnes dans l'ordre métier
            fusion = fusion[[c for c in ORDER if c in fusion.columns]
                            + [c for c in fusion.columns if c not in ORDER]]

            # ➤ Sauvegarde Excel
            fusion.to_excel(out, index=False)


            # Sauvegarder en Excel
            fusion.to_excel(out, index=False)

            # Appliquer le formatage Excel
            wb = load_workbook(out)
            ws = wb.active

            # Définir la plage du tableau et créer une table formatée
            last_col_letter = get_column_letter(ws.max_column)
            last_row = ws.max_row
            table_range = f"A1:{last_col_letter}{last_row}"

            # Créer et appliquer la table avec style
            table = Table(displayName="HistoriqueTable", ref=table_range)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            
            # Supprimer toute table existante et ajouter la nouvelle
            ws._tables.clear()
            ws.add_table(table)

            # Formater les colonnes spécifiques
            for idx, column in enumerate(ws[1], 1):
                col_letter = get_column_letter(idx)
                
                # Formater la colonne MONTH comme date
                if column.value == "MONTH":
                    for cell in ws[col_letter][1:]:  # Skip header
                        if cell.value:
                            try:
                                if isinstance(cell.value, (datetime, pd.Timestamp)):
                                    # Déjà en datetime, appliquer juste le format
                                    cell.number_format = "dd/mm/yyyy"
                                else:
                                    # Convertir en datetime Excel
                                    date_val = pd.to_datetime(cell.value, errors="coerce")
                                    if pd.notna(date_val):
                                        cell.value = date_val
                                        cell.number_format = "dd/mm/yyyy"
                            except Exception:
                                pass

                # Formater uniquement la colonne "TURNOVER €" avec le symbole €
                elif column.value == "TURNOVER €":
                    for cell in ws[col_letter][1:]:
                        if cell.value and isinstance(cell.value, (int, float)):
                            cell.number_format = "#,##0.00 €"

                # Formater les autres colonnes monétaires sans le symbole €
                elif column.value in ["TURNOVER", "C.A en €", "VARIABLE COSTS", "COGS", "VAR Margin", "Margin"]:
                    for cell in ws[col_letter][1:]:
                        if cell.value and isinstance(cell.value, (int, float)):
                            cell.number_format = "#,##0.00"

                # Formater la colonne QUANTITY
                elif column.value == "QUANTITY":
                    for cell in ws[col_letter][1:]:
                        if cell.value and isinstance(cell.value, (int, float)):
                            cell.number_format = "#,##0"

            # Figer la première ligne
            ws.freeze_panes = "A2"

            # Ajuster la largeur des colonnes
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = min(adjusted_width, 50)

            wb.save(out)
            self.txt_log_historique.appendPlainText(f"✅ Fusion terminée avec mise en forme optimisée. Fichier créé : {out}")
            self.pbar_historique.setValue(100)
        except Exception as e:
            self.txt_log_historique.appendPlainText(f"[ERROR] ❌ Une erreur est survenue pendant la fusion : {e}")
            import traceback
            traceback.print_exc()

    def _check_histo_columns_in_files(self):
        files = self.lst_historique_files.files()
        expected = self.histo_column_status_bar.expected_columns
        # On crée un mapping : nom attendu normalisé -> nom affiché
        expected_map = {col.strip().upper(): col for col in expected}
        presence = {col: set() for col in expected}

        for path in files:
            try:
                xls = pd.ExcelFile(path, engine="openpyxl")
                for sh in xls.sheet_names:
                    df = pd.read_excel(path, sheet_name=sh, nrows=0)
                    cols = [c.strip().upper() for c in df.columns]
                    for col_norm, col_affiche in expected_map.items():
                        if col_norm in cols:
                            presence[col_affiche].add(path)
            except Exception as e:
                self.txt_log_historique.appendPlainText(f"[WARN] ⚠ Fichier ignoré dans la détection : {path} – {e}")
        self.histo_column_status_bar.update_status_interactive(presence, files)

    # ---------- UI construction ----------
    def _build_traitement_ui(self, parent_widget):
        layout = QVBoxLayout(parent_widget)
        layout.setContentsMargins(18, 18, 18, 18)
        layout.setSpacing(12)
        
        # ► Sélecteur de date + bouton Charger taux
        row_date = QHBoxLayout()
        row_date.addWidget(QLabel("Date des taux :"))
        self.date_edit = QDateEdit(QDate.currentDate())
        self.date_edit.setCalendarPopup(True)
        row_date.addWidget(self.date_edit)
        btn_rates = QPushButton("Charger taux")
        btn_rates.clicked.connect(self._load_rates)
        row_date.addWidget(btn_rates)
        row_date.addStretch()
        layout.addLayout(row_date)

        # Taux manuel
        self.row_manual = QHBoxLayout()
        self.row_manual.addWidget(QLabel("Taux manuels (USD=0.93,GBP=1.15) :"))
        self.txt_manual = QLineEdit()
        self.row_manual.addWidget(self.txt_manual)
        layout.addLayout(self.row_manual)

        # Liste de fichiers
        layout.addWidget(QLabel("Fichiers Excel :"))
        self.lst_files = DropListWidget(on_click_callback=self._add_files)
        layout.addWidget(self.lst_files)

        # Boutons Ajouter / Retirer
        btn_bar = QHBoxLayout()
        btn_add = QPushButton("Ajouter…")
        btn_add.clicked.connect(self._add_files)
        btn_detect = QPushButton("🗓️ Détecter le ou les mois à traiter")
        btn_detect.clicked.connect(self._detect_months)
        btn_bar.addWidget(btn_detect)

        btn_rem = QPushButton("Retirer sélection")
        btn_rem.clicked.connect(self._remove_files)
        btn_bar.addWidget(btn_rem)
        btn_bar.addStretch()
        layout.addLayout(btn_bar)
        self.lst_files.setAlternatingRowColors(True)

        # Touche Suppr
        delete_act = QAction(
            self,
            shortcut=QKeySequence(Qt.Key.Key_Delete),
            triggered=self._remove_files
        )
        self.lst_files.addAction(delete_act)
        
        # Colonnes attendues - déplacé juste après la zone d'insertion de fichiers
        expected_columns = [
            "MONTH", "SIAMP UNIT", "SALE TYPE", "TYPE OF CANAL", "ENSEIGNE", 
            "CUSTOMER NAME", "COMMERCIAL AREA", "SUR FAMILLE", "FAMILLE", 
            "REFERENCE", "PRODUCT NAME", "QUANTITY", "TURNOVER", 
            "CURRENCY", "COUNTRY", "VARIABLE COSTS", "COGS"
        ]
        
        # Bandeau des colonnes avec défilement horizontal
        layout.addWidget(QLabel("<b>Colonnes attendues :</b>"))
        scroll_area, self.column_status_bar = ColumnStatusBar.create_scrollable(expected_columns)
        layout.addWidget(scroll_area)
        
        btn_check = QPushButton("Vérifier le contenu")
        btn_check.clicked.connect(self._check_columns_in_files)
        layout.addWidget(btn_check)

        # Chemin de sortie
        row_out = QHBoxLayout()
        row_out.addWidget(QLabel("Fichier de sortie :"))
        self.txt_out = QLineEdit("fusion.xlsx")
        btn_out = QPushButton("Parcourir…")
        btn_out.clicked.connect(self._choose_output)
        row_out.addWidget(self.txt_out)
        row_out.addWidget(btn_out)
        layout.addLayout(row_out)

        # Barre de progression
        self.pbar = QProgressBar()
        self.pbar.setMaximum(100)
        self.pbar.setValue(0)
        layout.addWidget(self.pbar)

        # Bouton Lancer
        btn_run = QPushButton("▶ Lancer")
        btn_run.setMinimumHeight(38)
        btn_run.clicked.connect(self._run_etl)
        layout.addWidget(btn_run)

        # Console intégrée
        self.txt_log = QPlainTextEdit()
        self.txt_log.setReadOnly(True)
        self.txt_log.setMaximumBlockCount(1000)
        layout.addWidget(self.txt_log, stretch=2)


    # ---------- style ----------
    def _apply_style(self):
        # Force le style sombre en définissant explicitement toutes les couleurs
        self.setStyleSheet("""
            QWidget { 
                font-family: 'Segoe UI', sans-serif; 
                font-size: 10pt; 
                color: #E0E0E0; 
                background-color: #22252A;
            }
            QMainWindow, QDialog, QTabWidget, QTabBar, QMenuBar, QMenu { 
                background-color: #22252A; 
                color: #E0E0E0;
            }
            QTabBar::tab {
                background-color: #2D3036;
                color: #E0E0E0;
                padding: 8px;
                margin-right: 2px;
            }
            QTabBar::tab:selected {
                background-color: #44576D;
                color: white;
            }
            QTabWidget::pane {
                border: 1px solid #444;
                background-color: #22252A;
            }
            QLabel { 
                font-weight: 500; 
                color: #E0E0E0; 
                background-color: transparent;
            }
            QLineEdit, QListWidget, QComboBox, QPlainTextEdit, QDateEdit { 
                background-color: #2D3036; 
                color: #E0E0E0;
                border: 1px solid #444; 
                padding: 4px; 
                border-radius: 4px; 
            }
            QComboBox::drop-down {
                border: none;
                background-color: #44576D;
            }
            QComboBox QAbstractItemView {
                background-color: #2D3036;
                color: #E0E0E0;
                selection-background-color: #44576D;
            }
            QPushButton { 
                background-color: #44576D; 
                color: #E0E0E0;
                border: none; 
                padding: 8px 12px; 
                border-radius: 4px; 
            }
            QPushButton:hover { 
                background-color: #527191; 
            }
            QPushButton:pressed { 
                background-color: #3C4E65; 
            }
            QListWidget { 
                border: 1px dashed #555; 
                background-color: #2D3036;
                color: #E0E0E0;
            }
            QScrollArea { 
                background-color: transparent;
                border: 1px solid #444;
                border-radius: 4px;
            }
            QScrollBar:horizontal {
                height: 10px;
                background: #2D3036;
                border-radius: 5px;
            }
            QScrollBar::handle:horizontal {
                background: #44576D;
                border-radius: 5px;
                min-width: 20px;
            }
            QScrollBar::handle:horizontal:hover {
                background: #527191;
            }
            QScrollBar:vertical {
                width: 10px;
                background: #2D3036;
                border-radius: 5px;
            }
            QScrollBar::handle:vertical {
                background: #44576D;
                border-radius: 5px;
                min-height: 20px;
            }
            QScrollBar::handle:vertical:hover {
                background: #527191;
            }
            QScrollBar::add-line, QScrollBar::sub-line {
                width: 0px;
                height: 0px;
            }
            QScrollBar::add-page, QScrollBar::sub-page {
                background: none;
            }
            QProgressBar {
                border: 1px solid #444;
                border-radius: 4px;
                background-color: #2D3036;
                color: #E0E0E0;
                text-align: center;
            }
            QProgressBar::chunk {
                background-color: #44576D;
                width: 10px;
            }
            QMessageBox {
                background-color: #22252A;
                color: #E0E0E0;
            }
            QMessageBox QLabel {
                color: #E0E0E0;
            }
            QMessageBox QPushButton {
                min-width: 80px;
            }
            QHeaderView::section {
                background-color: #44576D;
                color: #E0E0E0;
                padding: 4px;
                border: 1px solid #555;
            }
            QCalendarWidget {
                background-color: #2D3036;
                color: #E0E0E0;
            }
            QCalendarWidget QToolButton {
                color: #E0E0E0;
                background-color: #44576D;
                border: none;
            }
            QCalendarWidget QMenu {
                background-color: #2D3036;
                color: #E0E0E0;
            }
            QCalendarWidget QSpinBox {
                background-color: #2D3036;
                color: #E0E0E0;
                selection-background-color: #44576D;
            }
            QCalendarWidget QAbstractItemView {
                background-color: #22252A;
                color: #E0E0E0;
                selection-background-color: #44576D;
            }
            QToolTip {
                background-color: #2D3036;
                color: #E0E0E0;
                border: 1px solid #555;
            }
        """)

    @staticmethod
    def _iter_widgets(layout):
        return (layout.itemAt(i).widget() for i in range(layout.count()))

    def _add_files(self):
        files, _ = QFileDialog.getOpenFileNames(self, "Sélectionner fichiers", "", "Excel (*.xlsx)")
        for f in files:
            if f not in self.lst_files.files():
                self.lst_files.addItem(f)

    def _remove_files(self):
        for item in self.lst_files.selectedItems():
            self.lst_files.takeItem(self.lst_files.row(item))

    def _choose_output(self):
        path, _ = QFileDialog.getSaveFileName(self, "Fichier de sortie", self.txt_out.text(), "Excel (*.xlsx)")
        if path:
            self.txt_out.setText(path)

    def _run_etl(self):
        try:
            logging.info("Début de _run_etl")
            
            # DIAGNOSTIC: Vérifier l'existence et le contenu du script ETL_SIAMP.py
            self._check_script_integrity()
            
            files = self.lst_files.files()
            logging.info(f"Fichiers sélectionnés: {files}")
            
            # ➤ Détecter tous les mois distincts présents dans les fichiers
            from collections import defaultdict
            mois_detectés = defaultdict(list)

            for path in files:
                try:
                    logging.info(f"Analyse du fichier: {path}")
                    xls = pd.ExcelFile(path, engine="openpyxl")
                    for sh in xls.sheet_names:
                        df = xls.parse(sh, usecols="A:Q")
                        df.columns = [c.strip().upper() for c in df.columns]
                        if "MONTH" in df.columns:
                            mois = pd.to_datetime(df["MONTH"], errors="coerce").dt.to_period("M")
                            mois_uniques = sorted(mois.dropna().unique())
                            for m in mois_uniques:
                                mois_detectés[str(m)].append(os.path.basename(path))
                except Exception as e:
                    logging.exception(f"Erreur lors de l'analyse du fichier {path}")
                    self.txt_log.appendPlainText(f"[WARN] ⚠ Fichier ignoré : {path} – {e}")

            if not files:
                logging.warning("Aucun fichier sélectionné")
                return QMessageBox.warning(self, "Erreur", "Ajoutez au moins un fichier Excel.")
            
            out = self.txt_out.text().strip()
            logging.info(f"Chemin de sortie: {out}")
            
            if not out:
                logging.warning("Pas de chemin de sortie spécifié")
                return QMessageBox.warning(self, "Erreur", "Spécifiez le fichier de sortie.")

            man = self.txt_manual.text().strip()
            logging.info(f"Taux manuels: {man}")
            
            if not man:
                self.txt_log.appendPlainText("💡 Aucun taux manuel saisi. Le programme utilisera uniquement les taux ECB.")

            # Chemin du script embarqué
            base_path = getattr(sys, '_MEIPASS', os.path.abspath("."))
            script_path = os.path.join(base_path, "ETL_SIAMP.py")
            logging.info(f"Chemin du script: {script_path}")
            
            # Vérifier si le fichier ETL_SIAMP.py existe
            if not os.path.exists(script_path):
                logging.error(f"Le script {script_path} n'existe pas!")
                self.txt_log.appendPlainText(f"[ERREUR] Le script {script_path} n'existe pas!")
                # Essayer de trouver le script dans le répertoire courant
                current_dir_script = os.path.join(os.path.abspath("."), "ETL_SIAMP.py")
                if os.path.exists(current_dir_script):
                    logging.info(f"Script trouvé dans le répertoire courant: {current_dir_script}")
                    script_path = current_dir_script
                    self.txt_log.appendPlainText(f"[INFO] Script trouvé dans le répertoire courant.")
                else:
                    logging.error("Script introuvable dans le répertoire courant également!")
                    self.txt_log.appendPlainText("[ERREUR CRITIQUE] Script ETL_SIAMP.py introuvable!")
                    return QMessageBox.critical(self, "Erreur critique", "Le script ETL_SIAMP.py est introuvable!")

            # Essayer d'exécuter directement le script en mode intégré (sans subprocess)
            try:
                logging.info("Tentative d'exécution directe du script (mode intégré)")
                self.txt_log.appendPlainText("[INFO] Tentative d'exécution directe du script...")
                
                # Sauvegarder les arguments originaux
                original_argv = sys.argv.copy()
                
                # Préparer les arguments pour le script
                sys.argv = [script_path, "--chemin_sortie", out, "--fichiers", *files]
                if man:
                    sys.argv.extend(["--taux_manuels", man])
                date_str = self.date_edit.date().toString("yyyy-MM-dd")
                sys.argv.extend(["--date", date_str])
                if hasattr(self, "mois_selectionnes") and self.mois_selectionnes:
                    sys.argv.extend(["--mois_selectionnes", ",".join(self.mois_selectionnes)])
                
                logging.info(f"Arguments pour exécution directe: {sys.argv}")
                
                # Rediriger stdout vers notre console
                original_stdout = sys.stdout
                from io import StringIO
                captured_output = StringIO()
                
                class TeeOutput:
                    def __init__(self, console_callback, file_obj):
                        self.console_callback = console_callback
                        self.file_obj = file_obj
                    
                    def write(self, data):
                        self.console_callback(data)
                        self.file_obj.write(data)
                        return len(data)
                    
                    def flush(self):
                        self.file_obj.flush()
                
                sys.stdout = TeeOutput(
                    lambda x: self.txt_log.appendPlainText(x.rstrip()) if x.strip() else None,
                    captured_output
                )
                
                # Exécuter le script
                try:
                    # Importer le script comme module
                    import importlib.util
                    spec = importlib.util.spec_from_file_location("etl_module", script_path)
                    etl_module = importlib.util.module_from_spec(spec)
                    spec.loader.exec_module(etl_module)
                    
                    # Exécuter la fonction main
                    if hasattr(etl_module, 'main'):
                        etl_module.main()
                    else:
                        logging.error("Fonction main() non trouvée dans le script")
                        self.txt_log.appendPlainText("[ERREUR] Fonction main() non trouvée dans le script")
                    
                    self.txt_log.appendPlainText("[INFO] Exécution directe terminée avec succès")
                    self.pbar.setValue(100)
                    self._on_done(True)
                    return
                    
                except Exception as e:
                    logging.exception("Erreur lors de l'exécution directe du script")
                    self.txt_log.appendPlainText(f"[ERREUR] Exécution directe: {e}")
                    import traceback
                    self.txt_log.appendPlainText(f"[TRACEBACK]\n{traceback.format_exc()}")
                finally:
                    # Restaurer stdout et argv
                    sys.stdout = original_stdout
                    sys.argv = original_argv
                    
            except Exception as e:
                logging.exception("Erreur lors de la tentative d'exécution directe")
                self.txt_log.appendPlainText(f"[ERREUR] Tentative d'exécution directe: {e}")
            
            # Si l'exécution directe a échoué, revenir à l'exécution via subprocess
            self.txt_log.appendPlainText("[INFO] Utilisation de la méthode subprocess...")

            # Trouve python.exe (depuis PATH ou venv)
            python_exe = shutil.which("python") or sys.executable
            logging.info(f"Exécutable Python: {python_exe}")
            
            if not os.path.exists(python_exe):
                logging.error(f"L'exécutable Python {python_exe} n'existe pas!")
                self.txt_log.appendPlainText(f"[ERREUR] L'exécutable Python {python_exe} n'existe pas!")
                return QMessageBox.critical(self, "Erreur critique", f"L'exécutable Python {python_exe} est introuvable!")

            cmd = [python_exe, script_path, "--chemin_sortie", out, "--fichiers", *files]
            if man:
                cmd += ["--taux_manuels", man]
            date_str = self.date_edit.date().toString("yyyy-MM-dd")
            cmd += ["--date", date_str]
            if hasattr(self, "mois_selectionnes") and self.mois_selectionnes:
                cmd += ["--mois_selectionnes", ",".join(self.mois_selectionnes)]

            logging.info(f"Commande complète: {cmd}")
            env = dict(os.environ, GOOEY="0")
            logging.info("Environnement préparé")

            self.txt_log.clear()
            self.pbar.setValue(0)

            logging.info("Création du worker")
            self.worker = Worker(cmd, env)
            self.worker.log.connect(self.txt_log.appendPlainText)
            self.worker.progress.connect(self.pbar.setValue)
            self.worker.done.connect(self._on_done)
            
            logging.info("Démarrage du worker")
            self.worker.start()
            logging.info("Worker démarré avec succès")
            
        except Exception as e:
            logging.exception("Exception dans _run_etl")
            import traceback
            error_details = traceback.format_exc()
            self.txt_log.appendPlainText(f"[ERREUR] Une exception s'est produite : {e}\n\n{error_details}")
            QMessageBox.critical(self, "Erreur", f"Erreur détaillée : {e}\n\nConsultez le log pour plus de détails.")
            self.pbar.setValue(0)

    def _on_done(self, ok: bool):
        self.pbar.setValue(100 if ok else 0)
        logging.info(f"Traitement terminé avec statut: {'succès' if ok else 'échec'}")
        
        if ok:
            QMessageBox.information(
                self,
                "Terminé",
                "Traitement terminé avec succès !"
            )
        else:
            # Vérifier si un fichier d'erreur existe
            error_log_path = os.path.join(tempfile.gettempdir(), "etl_siamp_subprocess.log")
            app_log_path = os.path.join(tempfile.gettempdir(), f"etl_siamp_log_*.txt")
            
            error_details = ""
            if os.path.exists(error_log_path):
                try:
                    with open(error_log_path, "r", encoding="utf-8", errors="replace") as f:
                        error_content = f.read().strip()
                        if error_content:
                            error_details = f"\n\nDétails de l'erreur:\n{error_content}"
                except Exception as e:
                    logging.exception("Erreur lors de la lecture du fichier d'erreur")
                    error_details = f"\n\nImpossible de lire le fichier d'erreur: {e}"
            
            # Rechercher les fichiers de log de l'application
            import glob
            log_files = glob.glob(app_log_path)
            log_info = ""
            if log_files:
                log_info = f"\n\nLogs disponibles dans: {log_files[-1]}"
            
            QMessageBox.critical(
                self,
                "Erreur",
                f"Le script a échoué. Veuillez vérifier les fichiers d'entrée et les paramètres.{error_details}{log_info}\n\nConsultez la console pour plus de détails."
            )

    def _load_rates(self):
        try:
            from datetime import datetime, timedelta
            from ETL_SIAMP import get_ecb_rates

            date = self.date_edit.date().toString("yyyy-MM-dd")
            limit_date = (datetime.strptime(date, "%Y-%m-%d") - timedelta(days=60)).strftime("%Y-%m-%d")
            rates = get_ecb_rates(date)

            # ➕ Ajouter manuellement les devises non couvertes par l'ECB
            rates.update({
                "MAD": 0.094,
                "TND": 0.30,
                "DZD": 0.0068,
                "XOF": 0.0015
            })

            # 🔎 Analyser les fichiers chargés pour détecter les devises utilisées
            devises_utilisées = set()
            TURNOVER_SHEET = re.compile(r"^TURNOVER($|\s+[A-Z][a-z]{2}\s+\d{1,2}$)", re.I)
            for i in range(self.lst_files.count()):
                path = self.lst_files.item(i).text()
                try:
                    xls = pd.ExcelFile(path, engine="openpyxl")
                    for sh in filter(TURNOVER_SHEET.match, xls.sheet_names):
                        df = xls.parse(sh, usecols="A:Q")
                        df.columns = [str(c).strip().upper() for c in df.columns]
                        if "CURRENCY" in df.columns:
                            devises_utilisées.update(df["CURRENCY"].dropna().astype(str).str.strip().str.upper())
                except Exception as e:
                    self.txt_log.appendPlainText(f"[WARN] ⚠ Impossible de lire {path} : {e}")

            # 🖨️ Affichage dans la console de l'UI
            self.txt_log.appendPlainText(f"📅 Taux de change ECB au {date} :\n")

            taux_manuels = self.txt_manual.text().strip()
            manuels = dict(part.split("=") for part in taux_manuels.split(",") if "=" in part)
            manuels = {k.strip().upper(): float(v) for k, v in manuels.items()}
            
            if not devises_utilisées:
                self.txt_log.appendPlainText("[INFO] Aucune devise détectée dans les fichiers, veuillez glisser déposer vos fichiers à traiter pour détécter les devises.\n")
            else:
                for cur in sorted(devises_utilisées):
                    if cur in rates:
                        taux = rates[cur]
                        # Indiquer si le taux est direct ou inversé
                        if taux <= 1:
                            self.txt_log.appendPlainText(f"  • {cur:<4} → {taux:.6f} (1€ = {taux} {cur})")
                        else:
                            self.txt_log.appendPlainText(f"  • {cur:<4} → {taux:.6f} (1{cur} = {1/taux:.6f}€)")
                    elif cur in manuels:
                        taux = manuels[cur]
                        if taux <= 1:
                            self.txt_log.appendPlainText(f"  • {cur:<4} → {taux:.6f} (manuel, 1€ = {taux} {cur})")
                        else:
                            self.txt_log.appendPlainText(f"  • {cur:<4} → {taux:.6f} (manuel, 1{cur} = {1/taux:.6f}€)")
                    else:
                        val, ok = QInputDialog.getDouble(
                            self, f"Taux manquant pour {cur}",
                            f"Aucun taux trouvé pour {cur}.\nEntrez le taux de conversion vers EUR :",
                            min=0.0001, decimals=6
                        )
                        if ok:
                            manuels[cur] = val
                            if val <= 1:
                                self.txt_log.appendPlainText(f"  • {cur:<4} → {val:.6f} (ajouté manuellement, 1€ = {val} {cur})")
                            else:
                                self.txt_log.appendPlainText(f"  • {cur:<4} → {val:.6f} (ajouté manuellement, 1{cur} = {1/val:.6f}€)")
                        else:
                            self.txt_log.appendPlainText(f"  • {cur:<4} → ❌ Non disponible")

                # Mise à jour du champ texte
                self.txt_manual.setText(",".join(f"{k}={v}" for k, v in manuels.items()))

        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Erreur lors de la récupération ECB :\n{e}")


    def _build_parametres_ui(self, parent_widget):
        layout = QVBoxLayout(parent_widget)

        layout.addWidget(QLabel("Fichier de référence unique :"))

        # Zone de glisser-déposer pour le fichier de référence
        layout.addWidget(QLabel("Glissez votre fichier de référence ici ou cliquez pour le sélectionner :"))
        self.lst_reference_file = DropListWidget(on_click_callback=self._add_reference_file)
        self.lst_reference_file.setMaximumHeight(100)
        layout.addWidget(self.lst_reference_file)

        # Boutons pour le fichier de référence
        btn_ref_bar = QHBoxLayout()
        btn_add_ref = QPushButton("Ajouter fichier de référence…")
        btn_add_ref.clicked.connect(self._add_reference_file)
        btn_rem_ref = QPushButton("Retirer")
        btn_rem_ref.clicked.connect(self._remove_reference_file)
        btn_ref_bar.addWidget(btn_add_ref)
        btn_ref_bar.addWidget(btn_rem_ref)
        btn_ref_bar.addStretch()
        layout.addLayout(btn_ref_bar)

        # Informations sur la structure attendue
        layout.addWidget(QLabel("<b>Structure attendue du fichier de référence :</b>"))
        info_text = QPlainTextEdit()
        info_text.setMaximumHeight(150)
        info_text.setReadOnly(True)
        info_text.setPlainText("""Feuille "table" :
• Colonne B : PRODUCT NAME
• Colonne C : Surfamille ret  
• Colonne G : CONCAT NAME (ENSEIGNE + CUSTOMER NAME)
• Colonne H : Enseigne ret

Feuille "ZONE AFFECTATION" :
• Colonne A : PAYS
• Colonne E : COMMERCIAL AREA""")
        layout.addWidget(info_text)

        # Bouton Sauvegarder
        btn_save = QPushButton("💾 Sauvegarder le chemin")
        btn_save.clicked.connect(self._save_reference_paths)
        layout.addWidget(btn_save)

        # Charger si config existe
        self._load_reference_paths()

        # Bouton Vérifier les mises à jour
        btn_check_update = QPushButton("Vérifier les mises à jour")
        btn_check_update.clicked.connect(self.check_for_update_manual)
        layout.addWidget(btn_check_update)
        
        # Séparateur
        separator = QFrame()
        separator.setFrameShape(QFrame.Shape.HLine)
        separator.setFrameShadow(QFrame.Shadow.Sunken)
        layout.addWidget(separator)
        
        # Section diagnostic
        layout.addWidget(QLabel("<b>Outils de diagnostic :</b>"))
        
        btn_diag = QPushButton("🔍 Diagnostiquer l'environnement")
        btn_diag.clicked.connect(self._run_diagnostics)
        layout.addWidget(btn_diag)
        
        # Console de diagnostic
        self.txt_diag = QPlainTextEdit()
        self.txt_diag.setReadOnly(True)
        self.txt_diag.setMaximumBlockCount(1000)
        layout.addWidget(self.txt_diag, stretch=2)

    def _add_reference_file(self):
        file, _ = QFileDialog.getOpenFileName(self, "Sélectionner fichier de référence", "", "Excel (*.xlsx)")
        if file:
            # Vérifier que le fichier contient les feuilles requises
            try:
                xls = pd.ExcelFile(file, engine="openpyxl")
                if "table" in xls.sheet_names and "ZONE AFFECTATION" in xls.sheet_names:
                    # Vider la liste et ajouter le nouveau fichier
                    self.lst_reference_file.clear()
                    self.lst_reference_file.addItem(file)
                    QMessageBox.information(self, "✅ Succès", f"Fichier de référence validé :\n- Feuille 'table' détectée\n- Feuille 'ZONE AFFECTATION' détectée")
                else:
                    missing_sheets = []
                    if "table" not in xls.sheet_names:
                        missing_sheets.append("table")
                    if "ZONE AFFECTATION" not in xls.sheet_names:
                        missing_sheets.append("ZONE AFFECTATION")
                    QMessageBox.warning(self, "Erreur", f"Feuilles manquantes : {', '.join(missing_sheets)}")
            except Exception as e:
                QMessageBox.critical(self, "Erreur", f"Impossible d'ouvrir le fichier :\n{e}")

    def _remove_reference_file(self):
        self.lst_reference_file.clear()

    def _save_reference_paths(self):
        config = configparser.ConfigParser()
        
        # Récupérer le chemin du fichier de référence
        reference_file = ""
        if self.lst_reference_file.count() > 0:
            reference_file = self.lst_reference_file.item(0).text()
        
        config['REFERENCES'] = {
            'reference_file': reference_file
        }
        
        # Créer le répertoire mydata s'il n'existe pas
        config_dir = os.path.dirname(CONFIG_REF_FILE)
        os.makedirs(config_dir, exist_ok=True)
        
        with open(CONFIG_REF_FILE, 'w') as cfgfile:
            config.write(cfgfile)
        QMessageBox.information(self, "Succès", "Le chemin du fichier de référence a été sauvegardé.")
        self._load_reference_paths()  # Recharge directement après sauvegarde

    def _load_reference_paths(self):
        if os.path.exists(CONFIG_REF_FILE):
            config = configparser.ConfigParser()
            config.read(CONFIG_REF_FILE)
            refs = config['REFERENCES']
            reference_file = refs.get('reference_file', '')

            # Vider la liste et ajouter le fichier de référence
            self.lst_reference_file.clear()
            if reference_file and os.path.exists(reference_file):
                self.lst_reference_file.addItem(reference_file)

            # ✅ Check si le fichier existe physiquement
            if reference_file and not os.path.exists(reference_file):
                QMessageBox.warning(self, "Attention", f"⚠️ Fichier de référence manquant ou invalide :\n{reference_file}")

    def _check_columns_in_files(self):
        files = self.lst_files.files()
        expected = self.column_status_bar.expected_columns
        # Dictionnaire : colonne -> set des fichiers où elle est présente
        presence = {col: set() for col in expected}
        for path in files:
            try:
                xls = pd.ExcelFile(path, engine="openpyxl")
                for sh in xls.sheet_names:
                    df = pd.read_excel(path, sheet_name=sh, nrows=0)
                    cols = [c.strip().upper() for c in df.columns]
                    for col in expected:
                        if col in cols:
                            presence[col].add(path)
            except Exception as e:
                self.txt_log.appendPlainText(f"[WARN] ⚠ Fichier ignoré dans la détection : {path} – {e}")
        self.column_status_bar.update_status_interactive(presence, files)

    def check_for_update(self):
        try:
            # 1. Lire la version locale
            with open("version.txt", "r") as f:
                local_version = f.read().strip()

            # 2. Récupérer la version distante (exemple : fichier version.txt sur GitHub)
            url = "https://raw.githubusercontent.com/EliasGhennam/ETL_SIAMP/main/version.txt"
            response = requests.get(url, timeout=5)
            if response.status_code == 200:
                remote_version = response.text.strip()
                if remote_version != local_version:
                    self._show_update_dialog(remote_version)
        except Exception as e:
            print(f"Erreur lors de la vérification de mise à jour : {e}")

    def _show_update_dialog(self, remote_version, download_url):
        from PyQt6.QtWidgets import QMessageBox
        msg = QMessageBox(self)
        msg.setIcon(QMessageBox.Icon.Information)
        msg.setWindowTitle("Mise à jour disponible")
        msg.setText(f"Une nouvelle version ({remote_version}) est disponible !")
        msg.setInformativeText("Voulez-vous mettre à jour maintenant ?")
        msg.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        ret = msg.exec()
        if ret == QMessageBox.StandardButton.Yes:
            # Chemin du nouvel exécutable
            new_exe = "ETL_SIAMP_GUI_new.exe"
            if self.download_new_version(download_url, new_exe):
                # Préparer le script de remplacement
                with open("update.bat", "w") as f:
                    f.write(f"""
@echo off
timeout /t 2
move /y "{new_exe}" "{sys.argv[0]}"
start "" "{sys.argv[0]}"
""")
                QMessageBox.information(self, "Mise à jour", "L'application va se fermer pour mettre à jour.")
                os.startfile("update.bat")
                sys.exit(0)

    def download_new_version(self, download_url, new_filename):
        try:
            # Télécharge le nouvel exécutable
            response = requests.get(download_url, stream=True)
            with open(new_filename, "wb") as f:
                for chunk in response.iter_content(chunk_size=8192):
                    f.write(chunk)
            return True
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Erreur lors du téléchargement : {e}")
            return False

    def check_for_update_manual(self):
        try:
            with open("version.txt", "r") as f:
                local_version = f.read().strip()
            url = "https://raw.githubusercontent.com/EliasGhennam/ETL_SIAMP/main/version.txt"
            response = requests.get(url, timeout=5)
            if response.status_code == 200:
                remote_version = response.text.strip()
                if remote_version != local_version:
                    # Met à jour le lien de téléchargement si besoin
                    download_url = "https://github.com/EliasGhennam/ETL_SIAMP/releases/latest/download/ETL_SIAMP_GUI.exe"
                    self._show_update_dialog(remote_version, download_url)
                else:
                    QMessageBox.information(self, "Mise à jour", "Votre application est à jour.")
            else:
                QMessageBox.warning(self, "Erreur", "Impossible de vérifier la version distante.")
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Erreur lors de la vérification : {e}")

    def _check_script_integrity(self):
        """Fonction de diagnostic pour vérifier l'existence et le contenu du script ETL_SIAMP.py"""
        try:
            # Vérifier dans le répertoire PyInstaller
            base_path = getattr(sys, '_MEIPASS', os.path.abspath("."))
            script_path = os.path.join(base_path, "ETL_SIAMP.py")
            
            # Vérifier si le fichier existe
            if os.path.exists(script_path):
                logging.info(f"Script trouvé: {script_path}")
                self.txt_log.appendPlainText(f"[INFO] Script trouvé: {script_path}")
                
                # Vérifier la taille du fichier
                size = os.path.getsize(script_path)
                logging.info(f"Taille du script: {size} octets")
                self.txt_log.appendPlainText(f"[INFO] Taille du script: {size} octets")
                
                # Lire les premières lignes du script pour vérifier son contenu
                with open(script_path, 'r', encoding='utf-8', errors='replace') as f:
                    first_lines = ''.join(f.readlines(10))
                    logging.info(f"Début du script:\n{first_lines}")
                    self.txt_log.appendPlainText(f"[INFO] Début du script:\n{first_lines}")
                
                # Vérifier si le script est exécutable
                if not os.access(script_path, os.X_OK) and sys.platform != 'win32':
                    logging.warning(f"Le script n'est pas exécutable")
                    self.txt_log.appendPlainText(f"[WARN] Le script n'est pas exécutable")
                    os.chmod(script_path, 0o755)
                    logging.info(f"Permissions du script modifiées")
                    self.txt_log.appendPlainText(f"[INFO] Permissions du script modifiées")
            else:
                logging.error(f"Script non trouvé: {script_path}")
                self.txt_log.appendPlainText(f"[ERREUR] Script non trouvé: {script_path}")
                
                # Chercher dans le répertoire courant
                current_path = os.path.join(os.path.abspath("."), "ETL_SIAMP.py")
                if os.path.exists(current_path):
                    logging.info(f"Script trouvé dans le répertoire courant: {current_path}")
                    self.txt_log.appendPlainText(f"[INFO] Script trouvé dans le répertoire courant: {current_path}")
                    
                    # Copier le script dans le répertoire temporaire
                    import tempfile
                    temp_dir = tempfile.gettempdir()
                    temp_script = os.path.join(temp_dir, "ETL_SIAMP.py")
                    import shutil
                    shutil.copy2(current_path, temp_script)
                    logging.info(f"Script copié vers: {temp_script}")
                    self.txt_log.appendPlainText(f"[INFO] Script copié vers: {temp_script}")
                else:
                    # Rechercher le script dans tous les répertoires
                    import glob
                    potential_scripts = glob.glob("**/ETL_SIAMP.py", recursive=True)
                    if potential_scripts:
                        logging.info(f"Scripts potentiels trouvés: {potential_scripts}")
                        self.txt_log.appendPlainText(f"[INFO] Scripts potentiels trouvés: {potential_scripts}")
                    else:
                        logging.error("Aucun script ETL_SIAMP.py trouvé dans le système de fichiers")
                        self.txt_log.appendPlainText("[ERREUR] Aucun script ETL_SIAMP.py trouvé")
        except Exception as e:
            logging.exception("Erreur lors de la vérification du script")
            self.txt_log.appendPlainText(f"[ERREUR] Vérification du script: {e}")

    def _run_diagnostics(self):
        """Exécute un diagnostic complet de l'environnement et des dépendances"""
        self.txt_diag.clear()
        self.txt_diag.appendPlainText("=== DIAGNOSTIC DE L'ENVIRONNEMENT ===\n")
        
        try:
            # Informations système
            self.txt_diag.appendPlainText("--- INFORMATIONS SYSTÈME ---")
            self.txt_diag.appendPlainText(f"Système d'exploitation: {sys.platform}")
            self.txt_diag.appendPlainText(f"Version Python: {sys.version}")
            self.txt_diag.appendPlainText(f"Exécutable Python: {sys.executable}")
            self.txt_diag.appendPlainText(f"Répertoire courant: {os.getcwd()}")
            
            # Vérifier si on est dans un environnement PyInstaller
            if hasattr(sys, '_MEIPASS'):
                self.txt_diag.appendPlainText(f"Répertoire PyInstaller: {sys._MEIPASS}")
            else:
                self.txt_diag.appendPlainText("Pas dans un environnement PyInstaller")
            
            # Vérifier les variables d'environnement importantes
            self.txt_diag.appendPlainText("\n--- VARIABLES D'ENVIRONNEMENT ---")
            for var in ['PATH', 'PYTHONPATH', 'TEMP', 'TMP']:
                self.txt_diag.appendPlainText(f"{var}: {os.environ.get(var, 'Non défini')}")
            
            # Vérifier les dépendances
            self.txt_diag.appendPlainText("\n--- DÉPENDANCES ---")
            deps = {
                'pandas': pd.__version__ if 'pd' in globals() else "Non importé",
                'openpyxl': load_workbook.__module__ if 'load_workbook' in globals() else "Non importé",
                'PyQt6': Qt.__module__ if 'Qt' in globals() else "Non importé",
                'requests': requests.__version__ if 'requests' in globals() else "Non importé",
                'numpy': "Vérification...",
                'xml.etree.ElementTree': ET.__name__ if 'ET' in globals() else "Non importé"
            }
            
            # Vérifier numpy
            try:
                import numpy as np
                deps['numpy'] = np.__version__
            except ImportError:
                deps['numpy'] = "Non installé"
            
            for dep, ver in deps.items():
                self.txt_diag.appendPlainText(f"{dep}: {ver}")
            
            # Vérifier le script ETL_SIAMP.py
            self.txt_diag.appendPlainText("\n--- SCRIPT ETL_SIAMP.PY ---")
            base_path = getattr(sys, '_MEIPASS', os.path.abspath("."))
            script_path = os.path.join(base_path, "ETL_SIAMP.py")
            
            if os.path.exists(script_path):
                size = os.path.getsize(script_path)
                self.txt_diag.appendPlainText(f"Chemin: {script_path}")
                self.txt_diag.appendPlainText(f"Taille: {size} octets")
                self.txt_diag.appendPlainText(f"Dernière modification: {datetime.fromtimestamp(os.path.getmtime(script_path))}")
                
                # Vérifier le contenu du script
                try:
                    with open(script_path, 'r', encoding='utf-8', errors='replace') as f:
                        content = f.read(500)  # Lire les 500 premiers caractères
                        self.txt_diag.appendPlainText(f"\nDébut du fichier:\n{content}...")
                except Exception as e:
                    self.txt_diag.appendPlainText(f"Erreur lors de la lecture du script: {e}")
            else:
                self.txt_diag.appendPlainText(f"Script non trouvé: {script_path}")
                
                # Chercher dans d'autres emplacements
                locations = [
                    os.path.abspath("."),
                    os.path.join(os.path.dirname(sys.executable)),
                    os.path.expanduser("~"),
                    tempfile.gettempdir()
                ]
                
                for loc in locations:
                    path = os.path.join(loc, "ETL_SIAMP.py")
                    if os.path.exists(path):
                        self.txt_diag.appendPlainText(f"Script trouvé dans: {path}")
            
            # Vérifier les fichiers de log
            self.txt_diag.appendPlainText("\n--- FICHIERS DE LOG ---")
            log_pattern = os.path.join(tempfile.gettempdir(), "etl_siamp_*.log")
            subprocess_log = os.path.join(tempfile.gettempdir(), "etl_siamp_subprocess.log")
            
            import glob
            log_files = glob.glob(log_pattern)
            if log_files:
                self.txt_diag.appendPlainText(f"Logs trouvés: {log_files}")
                # Lire le dernier log
                latest_log = max(log_files, key=os.path.getmtime)
                try:
                    with open(latest_log, 'r', encoding='utf-8', errors='replace') as f:
                        content = f.read(1000)  # Lire les 1000 premiers caractères
                        self.txt_diag.appendPlainText(f"\nContenu du dernier log ({latest_log}):\n{content}...")
                except Exception as e:
                    self.txt_diag.appendPlainText(f"Erreur lors de la lecture du log: {e}")
            else:
                self.txt_diag.appendPlainText("Aucun fichier de log trouvé")
            
            # Vérifier le log du subprocess
            if os.path.exists(subprocess_log):
                try:
                    with open(subprocess_log, 'r', encoding='utf-8', errors='replace') as f:
                        content = f.read(1000)  # Lire les 1000 premiers caractères
                        self.txt_diag.appendPlainText(f"\nContenu du log subprocess:\n{content}...")
                except Exception as e:
                    self.txt_diag.appendPlainText(f"Erreur lors de la lecture du log subprocess: {e}")
            else:
                self.txt_diag.appendPlainText("Aucun log de subprocess trouvé")
            
            self.txt_diag.appendPlainText("\n=== DIAGNOSTIC TERMINÉ ===")
            
        except Exception as e:
            import traceback
            self.txt_diag.appendPlainText(f"ERREUR DURANT LE DIAGNOSTIC: {e}")
            self.txt_diag.appendPlainText(traceback.format_exc())


# --------------------------------------------------
# Lancement de l'application
# --------------------------------------------------
if __name__ == "__main__":
    app = QApplication(sys.argv)
    if hasattr(Qt.ApplicationAttribute, "AA_EnableHighDpiScaling"):
        app.setAttribute(Qt.ApplicationAttribute.AA_EnableHighDpiScaling)
    win = MainWindow()
    win.show()
    sys.exit(app.exec())

def normalize_columns(df):
    normalized_columns = df.columns.copy()
    for standard_name, variations in COLUMN_MAPPING.items():
        for col in df.columns:
            if col.strip().upper() in [v.strip().upper() for v in variations]:
                normalized_columns = normalized_columns.str.replace(col, standard_name)
    df.columns = normalized_columns
    return df

def clean_and_validate_data(df):
    # Conversion des dates
    if "MONTH" in df.columns:
        df["MONTH"] = pd.to_datetime(df["MONTH"], errors="coerce")
    
    # Nettoyage des valeurs numériques
    numeric_columns = ["TURNOVER", "QUANTITY"]
    for col in numeric_columns:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    
    # Standardisation des devises
    if "CURRENCY" in df.columns:
        df["CURRENCY"] = df["CURRENCY"].str.strip().str.upper()
    
    return df

def identify_merge_keys(df):
    potential_keys = ["REFERENCE", "CUSTOMER NAME", "MONTH"]
    available_keys = [key for key in potential_keys if key in df.columns]
    return available_keys

def format_month_column(df, year=None):
    """Formate correctement la colonne MONTH en gérant différents formats."""
    if "MONTH" not in df.columns:
        return df

    # Si valeurs entre 1-12 et année fournie => convertir en date complète
    numeric_months = pd.to_numeric(df["MONTH"], errors="coerce")
    month_mask = numeric_months.between(1, 12)
    if year and month_mask.any():
        df.loc[month_mask, "MONTH"] = pd.to_datetime(
            [f"{year}-{int(m):02d}-01" for m in numeric_months[month_mask]]
        )

    # Essai de conversion complète en date
    df["MONTH"] = pd.to_datetime(df["MONTH"], errors="coerce", dayfirst=True)
    
    return df

#pyinstaller --noconfirm --onefile --windowed --icon=mydata/siamp_icon.ico --add-data="ETL_SIAMP.py;." --add-data="mydata/*;mydata/" --name="ETL_SIAMP_GUI" ETL_SIAMP_GUI.py
#python -m PyInstaller --noconfirm ETL_SIAMP_GUI.spec
#python -m PyInstaller --noconfirm ETL_SIAMP_GUI_DEBUG.spec