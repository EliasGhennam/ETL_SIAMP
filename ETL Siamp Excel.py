import os
import pandas as pd
import re

# === Config ===
DOSSIER_SOURCE = "fichiers_excel"
DOSSIER_SORTIE = "output"
FICHIER_FINAL = "fusion_finale.xlsx"

# === Création des dossiers s'ils n'existent pas ===
os.makedirs(DOSSIER_SOURCE, exist_ok=True)
os.makedirs(DOSSIER_SORTIE, exist_ok=True)

# === Récupération des fichiers Excel ===
fichiers = [
    f for f in os.listdir(DOSSIER_SOURCE)
    if f.endswith('.xlsx') and not f.startswith('~$')  # On ignore les fichiers temporaires Excel
]

if not fichiers:
    print(f"Aucun fichier .xlsx trouvé dans le dossier '{DOSSIER_SOURCE}'.")
    exit()

print(f"{len(fichiers)} fichier(s) trouvé(s). Analyse des feuilles...")

dfs = []

# === Expression régulière pour "Turnover Oct 24" etc ===
pattern_turnover_date = re.compile(r"^Turnover\s+[A-Z][a-z]{2}\s+\d{1,2}$")

for fichier in fichiers:
    chemin = os.path.join(DOSSIER_SOURCE, fichier)
    try:
        xls = pd.ExcelFile(chemin, engine="openpyxl")
        feuilles_eligibles = [
            feuille for feuille in xls.sheet_names
            if feuille.lower() == "turnover" or pattern_turnover_date.match(feuille)
        ]

        if not feuilles_eligibles:
            print(f"❌ Aucune feuille 'Turnover' trouvée dans {fichier}. Ignoré.")
            continue

        for feuille in feuilles_eligibles:
            df = xls.parse(feuille, usecols="A:O")

            # === Harmonisation des colonnes ===
            # Unifie CURRENCY et Currency -> "Currency"
            if "CURRENCY" in df.columns and "Currency" not in df.columns:
                df.rename(columns={"CURRENCY": "Currency"}, inplace=True)
            elif "Currency" in df.columns and "CURRENCY" in df.columns:
                df["Currency"] = df["Currency"].combine_first(df["CURRENCY"])
                df.drop(columns=["CURRENCY"], inplace=True)

            # Unifie CUSTOMER NAME et Customer -> "Customer Name"
            if "CUSTOMER NAME" in df.columns and "Customer" not in df.columns:
                df.rename(columns={"CUSTOMER NAME": "Customer Name"}, inplace=True)
            elif "Customer" in df.columns and "CUSTOMER NAME" in df.columns:
                df["Customer Name"] = df["Customer"].combine_first(df["CUSTOMER NAME"])
                df.drop(columns=["Customer", "CUSTOMER NAME"], inplace=True)
            elif "Customer" in df.columns:
                df.rename(columns={"Customer": "Customer Name"}, inplace=True)

            # === Ajout de la source
            df["NomFichier"] = fichier
            df["Feuille"] = feuille
            dfs.append(df)


            print(f"✅ Feuille prise en compte : {feuille} ({fichier})")

    except ImportError as ie:
        print(f"❌ Erreur dans {fichier} : {ie}. Installe 'openpyxl' avec : pip install openpyxl")
    except Exception as e:
        print(f"❌ Erreur dans {fichier} : {e}")

# === Fusion finale ===
if not dfs:
    print("❌ Aucun fichier ou feuille valide n’a été trouvé. Arrêt.")
    exit()

fusion = pd.concat(dfs, ignore_index=True)

# === Option de tri (par exemple, par la première colonne si elle est une date ou nom) ===
# fusion = fusion.sort_values(by=fusion.columns[0])  # Active si nécessaire

# === Sauvegarde ===
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

# === Sauvegarde temporaire du fichier Excel (sans filtres)
chemin_final = os.path.join(DOSSIER_SORTIE, FICHIER_FINAL)
fusion.to_excel(chemin_final, index=False)

# === Réouverture avec openpyxl pour ajouter les filtres
wb = load_workbook(chemin_final)
ws = wb.active

# Calcul de la zone du tableau (A1 jusqu'à dernière colonne/ligne)
max_col = ws.max_column
max_row = ws.max_row
end_col_letter = chr(64 + max_col)  # 65 = A, 66 = B, etc.

table_range = f"A1:{end_col_letter}{max_row}"
table = Table(displayName="FusionTable", ref=table_range)

# Style du tableau (avec filtres activés automatiquement)
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
table.tableStyleInfo = style
ws.add_table(table)

wb.save(chemin_final)

print(f"\n✅ Fichier sauvegardé avec filtres Excel activés : {chemin_final}")
